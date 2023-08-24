import openpyxl
import re


def tech_data_get(exel_file: str, model_list: list = None,
                  exclusion_list: tuple = ('Интерполяция М', 'гофрирование', 'Интерполяция R')):
    """
    Функция получает технологические данные из файла exel_file технологических операций по списку моделей model_list,
    который равен списку имен листов файла исключая имена листов в списке exclusion_list
    :param exel_file: имя файла excel
    :param model_list: список моделей для получения данных - список листов книги excel
    :param exclusion_list: список листов книги excel файла для исключения из чтения
    :return:
    """
    # exclusion_list = ('Интерполяция М', 'гофрирование', 'Интерполяция R')
    ex_wb = openpyxl.load_workbook(exel_file, data_only=True)
    # получение списка моделей, если не указана модель
    # Если список моделей не указан явно, то забираем имена моделей из листов
    if model_list is None:
        model_list = []
        for sheet_name in ex_wb.sheetnames:
            if sheet_name not in exclusion_list:
                model_list.append(sheet_name)
    result_list = []  # список результатов
    for model_name in model_list:
        ex_sh = ex_wb[model_name]
        op_number_template = r'\d\d.\d\d*'  # шаблон номера операции
        # определение максимальной строки для чтения
        max_read_row = 1
        for row in ex_sh.iter_rows(min_row=2, min_col=1, max_row=ex_sh.max_row,
                                   max_col=2, values_only=True):
            max_read_row += 1
            if 'итого' in str(row[1]).lower():
                break
        # формирование списка результатов
        for row in ex_sh.iter_rows(min_row=1, min_col=1, max_row=max_read_row,
                                   max_col=12, values_only=True):
            if re.fullmatch(op_number_template, str(row[0])):
                result_list.append((model_name,  # модель котла
                                    row[0],  # номер операции
                                    row[1],  # имя операции
                                    row[2],  # имя РАЦ
                                    row[1] + '-' + row[2],  # полное имя операции
                                    str(row[3]),  # номер рц
                                    row[11]))  # расчётная трудоемкость
                previous_op_number = str(row[0])  # обработка объединенных ячеек имени операции
                previous_op_name = row[1]
                # print(row[0], '---', row[1], row[2], row[11], )
            elif row[0] is None and row[2] is not None:  # если объединённая ячейка
                result_list.append((model_name,  # модель котла
                                    previous_op_number,  # номер операции
                                    previous_op_name,  # имя операции
                                    row[2],  # имя рц
                                    previous_op_name + '-' + row[2],  # полное имя операции
                                    str(row[3]),  # номер рц
                                    row[11]))  # расчётная трудоемкость
    return model_list, result_list


if __name__ == '__main__':
    ex_file_dir_tst = r'D:\АСУП\Python\Projects\OmzitTerminal\Трудоёмкость серия I.xlsx'
    model_list_tst = ['7000М+', '800М+']
    exclusion_list_tst = ('Интерполяция М', 'гофрирование', 'Интерполяция R')
    tech_data_get(exel_file=ex_file_dir_tst)
