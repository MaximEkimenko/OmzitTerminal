"""
TODO создать рабочее место технолога для занесения и корректировки
Занесение исходных данных для стратегического планирования
"""
import json
from decimal import Decimal
from pathlib import Path
from pprint import pprint
from django.core.exceptions import ObjectDoesNotExist

import openpyxl

try:  # для запуска файла
    from scheduler.models import SeriesParameters, ModelParameters  # noqa
    from m_logger_settings import logger  # noqa
except Exception:
    pass


def series_parameters_set() -> None:
    """
    Первоначальное заполнение параметров серий
    :return: None
    """
    # полный по сериям массив данных
    new_series_parameters = [
        {
            'series_name': 'ML',
            'cycle_polynom_koef': list(reversed([Decimal('0.0003112364434866'),
                                                 Decimal('-0.0150453236513729'),
                                                 Decimal('0.201039119152367'),
                                                 Decimal('1.00653250263694'),
                                                 Decimal('16.2077210606424')])),
            'difficulty_koef': Decimal('1.3')
        },
        {
            'series_name': 'DMH',
            'cycle_polynom_koef': list(reversed([Decimal('-0.000130811023477673'),
                                                 Decimal('0.0118430669074363'),
                                                 Decimal('-0.348060948000969'),
                                                 Decimal('5.22023230341833'),
                                                 Decimal('3.67987872416067')])),
            'difficulty_koef': Decimal('1.5')

        },
        {
            'series_name': 'OV',
            'cycle_polynom_koef': list(reversed([Decimal('-0.000130811023477673'),
                                                 Decimal('0.0118430669074363'),
                                                 Decimal('-0.348060948000969'),
                                                 Decimal('5.22023230341833'),
                                                 Decimal('3.67987872416067')])),
            'difficulty_koef': Decimal('1.3')
        },
        {
            'series_name': 'SV',
            'cycle_polynom_koef': list(reversed([Decimal('-0.000130811023477673'),
                                                 Decimal('0.0118430669074363'),
                                                 Decimal('-0.348060948000969'),
                                                 Decimal('5.22023230341833'),
                                                 Decimal('3.67987872416067')])),
            'difficulty_koef': Decimal('1')
        },
        {
            'series_name': 'R',
            'cycle_polynom_koef': list(reversed([Decimal('0.21446424625724'),
                                                 Decimal('-2.00393257283599'),
                                                 Decimal('6.08211796013855'),
                                                 Decimal('-3.82285906703149'),
                                                 Decimal('6.12988710720749')])),
            'difficulty_koef': Decimal('1')
        },
        {
            'series_name': 'P',
            'cycle_polynom_koef': list(reversed([Decimal('-0.0000763316486661341'),
                                                 Decimal('0.00623813864707741'),
                                                 Decimal('-0.164871150491832'),
                                                 Decimal('3.12920114698329'),
                                                 Decimal('12.6240783289121')])),
            'difficulty_koef': Decimal('1')
        },
        {
            'series_name': 'M',
            'cycle_polynom_koef': list(reversed([Decimal('0.0003112364434866'),
                                                 Decimal('-0.0150453236513729'),
                                                 Decimal('0.201039119152367'),
                                                 Decimal('1.00653250263694'),
                                                 Decimal('16.2077210606424')])),
            'difficulty_koef': Decimal('1')
        },
    ]
    # добавление/обновление БД
    for data in new_series_parameters:
        try:
            parameters = SeriesParameters.objects.get(series_name=data['series_name'])
            if (parameters.cycle_polynom_koef != data['cycle_polynom_koef']
                    or parameters.difficulty_koef != data['difficulty_koef']):
                parameters.cycle_polynom_koef = data['cycle_polynom_koef']
                parameters.difficulty_koef = data['difficulty_koef']
                parameters.save()
                logger.info(f'Параметры серии {parameters} были обновлёны.')
            else:
                logger.debug(f'В параметрах серии {parameters} изменений не обнаружено.')
        except ObjectDoesNotExist:
            SeriesParameters.objects.create(series_name=data['series_name'],
                                            cycle_polynom_koef=data['cycle_polynom_koef'],
                                            difficulty_koef=data['difficulty_koef'])
            logger.info(f'Параметры новой серии {data["series_name"]} были добавлены.')


def get_all_weights(xlsx_paths: list) -> dict[str]:
    """
    Получение json всех масс котлов из xlsx технологов
    :param xlsx_paths:
    :return:
    """
    all_weights_json = {}
    for xlsx_path in xlsx_paths:
        xlsx_files = []
        for execl_file in Path(xlsx_path).rglob("*.xlsx"):
            if '~' not in execl_file.name:
                xlsx_files.append(execl_file)

        for xlsx_file in xlsx_files:
            wb = openpyxl.load_workbook(xlsx_file, data_only=True)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                model_weight = ws["G2"].value
                # поиск суммарной трудоёмкости
                full_norm_tech = 0
                for i in range(1, 2000):
                    if ws[f"O{i}"].value == 'часов':
                        full_norm_tech = ws[f"O{i - 1}"].value
                        try:
                            full_norm_tech = round(float(full_norm_tech), 2)
                        except ValueError as e:
                            full_norm_tech = 0
                        break
                print(f"{full_norm_tech=}")
                if model_weight is not None and (type(model_weight) is float or type(model_weight) is int):
                    all_weights_json[sheet] = {'weight': model_weight, 'file': xlsx_file.name,
                                               'full_norm_tech': full_norm_tech}
                if model_weight is not None and (type(model_weight) is float or type(model_weight) is int):
                    all_weights_json[sheet] = {'weight': model_weight, 'file': xlsx_file.name,
                                               'full_norm_tech': full_norm_tech}
    json_file_to_save = r'D:\АСУП\Python\Projects\OmzitTerminal\misc\all_weights.json'
    with open(json_file_to_save, 'w') as json_file:
        json.dump(all_weights_json, json_file, ensure_ascii=False, indent=4)
    return all_weights_json


def clean_model_names(models_data: dict[str]) -> dict[str:float]:
    """
    Очистка имён модели от лишних символов
    :param models_data: данные json
    :return: {имя модели: вес модели}
    """
    translate_dict = {
        '+': '', 'М': 'M', 'М+': 'M', 'M+': 'M', 'Р': 'P', ' ': '', '1': '1', '2': '2', '3': '3', '4': '4', '5': '5',
        '6': '6', '7': '7', '8': '8', '9': '9', '0': '0', 'R': 'R', 'Д': 'D', 'А': 'A'
    }
    clean_dict = {}
    for model_name, model_data in models_data.items():
        model_weight = round(model_data['weight'], 2)
        model_full_norm_tech = model_data['full_norm_tech']
        if len(model_name.split()) > 1:
            new_model_name = f"{model_name.split()[0]}{model_name.split()[1]}"


            new_data = clean_dict.get(new_model_name, 0)
            if new_data != 0:
                max_weight_value = max(model_weight, new_data.get('model_weight', 0))
            else:
                max_weight_value = model_weight


            if 'SW' in model_name.split()[1] or 'SV' in model_name.split()[1] or 'RC' in model_name.split()[1]:
                clean_dict.update({new_model_name: {'model_weight': max_weight_value,
                                                    'full_norm_tech': model_full_norm_tech}})
            elif 'ML' in model_name.split()[0] and 'ML' in model_name.split()[1]:

                new_model_name_left = f"{model_name.split()[0]}"
                new_model_name_right = f"{model_name.split()[1]}"[1:-1]

                clean_dict.update({new_model_name_left: {'model_weight': max_weight_value,
                                                         'full_norm_tech': model_full_norm_tech}})
                clean_dict.update({new_model_name_right: {'model_weight': max_weight_value,
                                                          'full_norm_tech': model_full_norm_tech}})

                pass
            else:
                latin_model_name = ''.join(translate_dict.get(char, char) for char in model_name.split()[0])
                new_model_name = latin_model_name


                # max_weight_value = max(model_weight, clean_dict.get(new_model_name, 0))

                new_data = clean_dict.get(new_model_name, 0)
                if new_data != 0:
                    max_weight_value = max(model_weight, new_data.get('model_weight', 0))
                else:
                    max_weight_value = model_weight


                clean_dict.update({new_model_name: {'model_weight': max_weight_value,
                                                    'full_norm_tech': model_full_norm_tech}})
        else:
            latin_model_name = ''.join(translate_dict.get(char, char) for char in model_name)
            clean_dict.update({latin_model_name: {'model_weight': model_weight,
                                                  'full_norm_tech': model_full_norm_tech}})
    return clean_dict


def models_data_db_set(data: dict) -> None:
    """
    Первоначальное заполнение параметров моделей
    :param data:
    :return:
    """
    series = ('ML', 'DMH', 'OV', 'SV', 'R', 'P', 'M')
    for model, models_data in data.items():
        model_name = model.split('-')[0]

        model_weight = Decimal(str(models_data['model_weight']))
        full_norm_tech = Decimal(str(models_data['full_norm_tech']))

        for series_element in series:
            if model_name.endswith(series_element):
                series_id = SeriesParameters.objects.get(series_name=series_element).id
                data_to_db = {'model_name': model_name,
                              'model_weight': model_weight,
                              'series_parameters_id': series_id,
                              'full_norm_tech': full_norm_tech}
                # print(data_to_db)
                params = ModelParameters(**data_to_db)
                params.save()


if __name__ == '__main__':
    xlsx_paths_tst = [
        r'M:\Xranenie\ПТО\котлы\Трудоемкость котлов\Трудоемкость котлов',
        r'M:\Xranenie\ПТО\котлы\Трудоемкость котлов\Трудоемкость экономайзеров',
        r'M:\Xranenie\ПТО\котлы\Трудоемкость котлов\Трудоемкость деаэраторов'
    ]
    # get_all_weights(xlsx_paths_tst)
    json_file_to_save_tst = r'D:\АСУП\Python\Projects\OmzitTerminal\misc\all_weights.json'
    with open(json_file_to_save_tst, 'r') as file:
        model_names_tst = json.load(file)
    # print(clean_model_names(model_names_tst))
    # models_data_db_set(clean_model_names(model_names_tst))
