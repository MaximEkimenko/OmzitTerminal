import json
import os
import re
import sys
import time

import psutil
import pythoncom
from win32com.client import Dispatch, gencache
import openpyxl as ox

import logging

logging.basicConfig(filename=os.path.join("cdw_reader.log"), filemode='a',
                    format='%(asctime)s - %(levelname)s - %(message)s',
                    level=logging.DEBUG)


class CDWReader:
    """
    Класс обработчика cdw чертежей для получения спецификаций.
    :parameter files_paths - Путь к файлу или список путей файлов чертежей
    :methods:
        create_json - Создание json файла общей спецификации
        create_xlsx - Создание xlsx файла общей спецификации
    """
    _module = gencache.EnsureModule(
        typelibCLSID="{69AC2981-37C0-4379-84FD-5DD2F3C0A520}",
        lcid=0,
        major=1,
        minor=0
    )
    _api = _module.IKompasAPIObject(
        Dispatch("Kompas.Application.7")._oleobj_.QueryInterface(
            _module.IKompasAPIObject.CLSID,
            pythoncom.IID_IDispatch
        )
    )
    _app = _api.Application  # Получаем основной интерфейс
    _const = gencache.EnsureModule(
        typelibCLSID="{75C9F5D0-B5B8-4526-8681-9903C567D2ED}",
        lcid=0,
        major=1,
        minor=0
    ).constants
    _title_correction = (
        ("длин", "Длина"),
        ("кол", "Количество"),
        ("прим", "Примечание"),
        ("наим", "Наименование"),
        ("обозн", "Обозначение"),
        ("мате", "Материал"),
        ("профиль", "Профиль"),
    )
    _material_pattern = r".*\$[a-z]+.+;.+\$"
    _material_split_pattern = r"(.*)\$[a-z]+(.+);(.+)\$"
    _profile_pattern = r"^@\d*~(.*)$"  # Профиль @134~120x120x4
    _patterns = (
        # Балка (200)\nШвеллер 30П $dmГОСТ 8240-97;С255-4 ГОСТ 27772-2015$
        (fr"^(.*)\((\d+)\)\s*\n+({_material_pattern})$", ("Наименование", "Длина", "Материал")),
        # Балка \n Швеллер 30П $dmГОСТ 8240-97;С255-4 ГОСТ 27772-2015$\nL=1200мм
        (fr"^(.*)\n+({_material_pattern})\n+L\s*=\s*([\d\.]+)\s*", ("Наименование", "Материал", "Длина")),
        # Балка \n Швеллер 30П $dmГОСТ 8240-97;С255-4 ГОСТ 27772-2015$
        (fr"^(.*)\n+({_material_pattern})$", ("Наименование", "Материал")),
        # Швеллер 30П $dmГОСТ 8240-97;С255-4 ГОСТ 27772-2015$
        (fr"^({_material_pattern})$", ("Материал",)),
        (fr"^(.*)\n+(.*ГОСТ.*)\n*$", ("Наименование", "Материал")),
        # (fr"^([а-яА-Я]+)\sL\s(\d+)", ("Наименование", "Длина")),
    )

    _columns = set()

    allowed_titles = ("Наименование", "Материал", "Количество", "Длина")
    _prohibited_texts = ("Детали", "Сборочные единицы", "Изготовить", "@", "Документация", "Стандартные изделия")

    def __init__(self, files_paths):
        self._app.Visible = False  # Скрываем окно от пользователя
        self._app.HideMessage = self._const.ksHideMessageNo  # Отвечаем НЕТ на любые вопросы программы

        self._files_paths = []
        # Проверяем, передан один или несколько файлов, а так же расширение cdw
        if isinstance(files_paths, str):
            if files_paths.endswith(".cdw"):
                self._files_paths.append(files_paths)
        elif hasattr(files_paths, "__iter__") and all(isinstance(path, str) for path in files_paths):
            cdw_files = [fp for fp in files_paths if fp.endswith(".cdw")]
            self._files_paths.extend(cdw_files)
        else:
            raise TypeError("Может быть передана только строка или коллекция строк!")

    def _open_document(self, path):
        logging.debug(f"Открытие документа {path}")
        document = self._app.Documents.Open(
            PathName=path,
            Visible=False,
            ReadOnly=True
        )
        logging.debug(f"Документ открыт {document}")
        return document

    def _close_all_documents(self):
        for document in self._app.Documents:
            logging.debug(f"Закрытие документа {document}")
            closed = document.Close(0)
            if closed:
                logging.debug(f"Документ закрыт!")
        if not self._app.Documents:
            logging.debug(f"Все документы закрыты!")

    def _close_kompas(self):
        logging.debug(f"Закрытие приложения {self._app}")
        self._app.Quit()
        if not self._app:
            logging.debug(f"Приложение закрыто!")

    def _get_all_data(self):
        """
        Собирает общую спецификацию из разных сценариев обработки
        :return:
            specification - Спецификация в виде словаря, первый ключ - columns - все названия столбцов
                            остальные ключи - имена файлов чертежей, содержащих список изделий в виде словарей,
                            где ключи - названия столбцов,
                            значения - соответствующий текст ячейки спецификации изделия
        """
        specification = dict()
        specification["columns"] = self.allowed_titles
        for file_path in self._files_paths:
            filename = os.path.split(file_path)[-1].replace(".cdw", "")
            specification[filename] = []
            document = self._open_document(file_path)
            try:
                for func in (self._parse_specifications, self._parse_tables):
                    data = func(document)
                    specification[filename].extend(data)
            except Exception as ex:
                logging.exception(ex)
        self._app.HideMessage = self._const.ksShowMessage
        self._close_all_documents()
        self._close_kompas()
        return specification

    def _parse_specifications(self, document):
        """
        Получает данные из таблиц чертежа, выполненных отдельными спецификациями, исключая скрытые
        :param document: Чертеж
        :return:
            specification - словарь, где ключи - названия столбцов,
                            значения - соответствующий текст ячейки спецификации изделия
        """
        specification = []
        tables = document.SpecificationDescriptions  # Не работает без лицензии !!!
        if tables:
            for table in tables:
                if table.ShowOnSheet:

                    rows = table.Objects
                    if rows:
                        for row in rows:
                            cells = row.Columns
                            if cells:
                                values = [cell.Text.Str for cell in cells]
                                titles = [cell.ColumnName for cell in cells]
                                self._columns.update(titles)
                                specification.append(dict(zip(titles, values)))
                        specification.append([])

        #  Объединяет обозначение детали с наименованием
        for i, row in enumerate(specification):
            if "Обозначение" in row and "Наименование" in row:
                x = row.pop("Обозначение")
                if x != "":
                    row["Наименование"] = x + " " + row["Наименование"]
                    if i != len(specification) - 1:
                        next_row = specification[i + 1]
                        if "Обозначение" in next_row and "Наименование" in next_row:
                            if next_row["Обозначение"] == "" and next_row["Наименование"]:
                                row["Наименование"] += "\n" + next_row["Наименование"]
                                specification[i + 1] = []
                else:
                    logging.debug(f"Третий случай: {x} строка {row}")
        specification = self._parsed_data_handler(specification)
        return specification

    def _parse_tables(self, document):
        """
        Получает данные из таблиц чертежа, выполненных угловыми спецификациями
        :param document: Чертеж
        :return:
            specification - словарь, где ключи - названия столбцов,
                            значения - соответствующий текст ячейки спецификации изделия
        """
        specification = []
        views = self._module.IKompasDocument2D(document).ViewsAndLayersManager.Views
        if views:
            for view in views:
                tables = self._module.IDrawingContainer(view).Objects(25)
                if tables:
                    for table in tables:
                        table = self._module.ITable(table)
                        col_count = table.ColumnsCount
                        table_width = 0
                        titles = []
                        for i_col in range(col_count):
                            cell = table.Cell(0, i_col)
                            table_width += self._module.ICellFormat(cell).Width
                            titles.append(self._module.IText(cell.Text).Str)
                            titles = self._titles_handler(titles)
                            if "Количество" not in titles and "Примечание" in titles:
                                titles[titles.index("Примечание") - 1] = "Количество"
                            self._columns.update(titles)
                        # Если ширина таблицы примерно равна ширине основной надписи
                        if 170 <= table_width <= 185:
                            row_count = table.RowsCount
                            for i_row in range(1, row_count):
                                values = [
                                    self._module.IText(table.Cell(i_row, i_col).Text).Str
                                    for i_col in range(col_count)
                                ]
                                specification.append(dict(zip(titles, values)))
                            specification.append([])
        specification = self._parsed_data_handler(specification)
        return specification

    def _parsed_data_handler(self, data):
        """
        Обработчик всех данных сформированных таблиц
        """
        if data:
            common_material = None
            for i_row, row in enumerate(data):
                # Если стока содержит не только пробелы или переносы
                if row and any((row and not re.findall(r'^[\s\n]*$', text) for text in row.values())):

                    if "Обозначение" in row:
                        # Случай, когда материал указан для нескольких деталей
                        if "Детали из" in row["Обозначение"]:
                            common_material = row["Наименование"]
                            data[i_row] = None
                        elif common_material:
                            row["Материал"] = common_material

                    # Убирает лишние символы из поля длина
                    if "Длина" in row:
                        row["Длина"] = row["Длина"].replace("-", "")

                    # Обработка таблиц с профилями
                    if "Профиль" in row:
                        profile = row.pop("Профиль")
                        match = re.match(self._profile_pattern, profile)
                        if match:
                            profile = match.group(1)
                        if profile not in ("-", "", " "):
                            name = row["Марка"] + " Профиль " + profile
                            row["Материал"] = row["Наименование"]
                            row["Наименование"] = name
                        elif "ГОСТ" in row["Наименование"]:
                            row["Материал"] = row["Наименование"]
                            row["Наименование"] = row["Марка"]
                        else:
                            logging.debug(f'Третий случай profile: {row}')

                    updates = dict()
                    for title, text in row.items():

                        for pattern in self._patterns:
                            match = re.match(pattern[0], text)
                            if match:
                                updates.update(dict(zip(pattern[1], match.groups())))
                                material_match = re.match(self._material_split_pattern, updates["Материал"])
                                if material_match:
                                    updates["Материал"] = " ".join(material_match.groups())
                                break
                        else:
                            for prohibited_text in self._prohibited_texts:
                                if prohibited_text in text:
                                    data[i_row] = []
                            row[title] = text.replace("\n", "")
                    row.update(updates)
                else:
                    common_material = None
                    data[i_row] = None
        unique_data = dict()
        for i, row in enumerate(data):
            if row:
                filtered_row = {}
                position_number = row.get("@", "")  # номер позиции
                # формируем уникальный ключ строки по разрешенным столбцам и номеру позиции
                row_key = ''.join((row.get(title, "") for title in self.allowed_titles)) + position_number
                if row_key not in unique_data:
                    # разбиваем строку по разрешенным столбцам
                    for k in self.allowed_titles:
                        value = row.get(k, "")
                        if k == "Наименование" and position_number.isdigit():
                            value = " ".join((position_number, value))  # добавляем номер позиции к наименованию
                        filtered_row[k] = value
                    unique_data[row_key] = filtered_row
                else:
                    # добавляем количество к уже существующей детали
                    unique_data[row_key]["Количество"] = str(sum(map(
                        int, (unique_data[row_key]["Количество"], row["Количество"])
                    )))
        filtered_data = list(unique_data.values())
        return filtered_data

    def _titles_handler(self, titles):
        """
        Обработчик заголовков таблиц в соответствии с _title_correction
        """
        for i, title in enumerate(titles):
            for bad, good in self._title_correction:
                if bad in title.lower():
                    titles[i] = good
        return titles

    def create_json(self, json_path):
        """
        Создает json файл спецификации
        :param json_path: Путь к json-файлу
        """
        spec = {}
        data = self._get_all_data()
        logging.debug(f"Создание JSON из данных {data}")
        if os.path.exists(json_path):
            # Проверяет, есть ли уже спецификация по данному пути,
            # если есть, значит она обновляется новыми данными (дозагрузка чертежей)
            try:
                with open(json_path, 'r') as json_file:
                    spec = json.load(json_file)
            except Exception as ex:
                logging.error(f"При чтении json файла вызвано исключение: {ex}")
        spec.update(data)
        with open(json_path, "w") as json_file:
            json.dump(spec, json_file)

    def create_xlsx(self, xlsx_path):
        """
        Создает файл Excel спецификации
        :param xlsx_path: Путь к файлу Excel
        """
        data = self._get_all_data()
        wb = ox.Workbook()
        ws = wb.worksheets[0]
        # Создаем шапку таблицы
        columns = data.pop("columns")
        for i_col, title in enumerate(columns):
            ws.cell(row=1, column=i_col + 1).value = title

        # Заполняем ячейки таблицы по ключам из заголовка соответствующего столбца
        row_number = 1
        for details in data.values():
            for i_row, row in enumerate(details):
                row_number += 1
                for i_col, key in enumerate(columns):
                    value = row.get(key)
                    if value:
                        ws.cell(row=row_number, column=i_col + 1).value = value

        wb.save(xlsx_path)


def get_specification(files):
    logging.info("Начало получения спецификации")
    start = time.time()
    try:
        logging.debug(f"Файлы: {files}")
        reader = CDWReader(files_paths=files)
        files_path = os.path.split(files[0])[0]
        reader.create_json(os.path.join(files_path, "specification.json"))
    except Exception as ex:
        logging.exception(ex)
        print(ex)
    finally:
        # Убиваем процессы КОМПАС
        for process in psutil.process_iter():
            if "KOMPAS" in process.name():
                process.kill()
    logging.info(f"Завершено! Время выполнения:{time.time() - start}")


if __name__ == "__main__":
    # Извлекаем из переданных аргументов список файлов
    if len(sys.argv) > 2:
        files = " ".join(sys.argv[1:])
    else:
        files = sys.argv[1]
    files = files.split("+-+")

    logging.debug(f"{files}")
    get_specification(files)
