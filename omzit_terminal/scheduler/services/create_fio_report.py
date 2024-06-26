import datetime
import json
import os
import time
from pprint import pprint

from scheduler.models import ShiftTask, Doers  # noqa
import openpyxl
from pathlib import Path
from django.shortcuts import render, redirect
from django.db.models import Q, Sum
import pandas
from omzit_terminal.settings import BASE_DIR  # noqa
from django.utils.timezone import make_aware, make_naive


def create_fio_report(fio: str, work_date: datetime.datetime) -> dict[str, tuple]:
    """
    Создание отчёта по fio в дату work_date
    :param fio: ФИО
    :param work_date: дата работы
    :return: dict из queryset
    """
    # директория хранения отчётов
    excel_save_path = Path.joinpath(BASE_DIR, 'daily_reports')
    # ежедневная очистка отчётов
    surname = fio.split()[0]
    name = fio.split()[1]
    work_date_start = make_aware(datetime.datetime(year=work_date.year, month=work_date.month,
                                        day=work_date.day, hour=0, minute=1))
    work_date_end = make_aware(datetime.datetime(year=work_date.year, month=work_date.month,
                                      day=work_date.day, hour=23, minute=59))
    queryset = (ShiftTask.objects.exclude(fio_doer="не распределено")
                .values('fio_doer', 'decision_time', 'norm_calc')
                .order_by("datetime_assign_wp")
                .filter(fio_doer__contains=fio)
                .filter(st_status='принято')
                .filter(Q(datetime_job_start__gte=work_date_start) & Q(datetime_job_start__lte=work_date_end))
                .values('fio_doer', 'decision_time', 'norm_calc', 'op_number', 'ws_number')
                )

    fio_data = {'fio_data': tuple(queryset)}
    # сохранение excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(('ФИО', 'ТД,ч', 'оп.№', 'T№', 'дата'))
    # print(queryset)
    if queryset:
        for line in queryset:
            row = (fio, line['norm_calc'], line['op_number'],  line['ws_number'],
                   line['decision_time'].strftime('%d.%m.%Y'))
            ws.append(row)
        ws.append((f"всего:{ws.max_row - 1}",  f'=SUM(B1:B{ws.max_row})'))
        excel_file_save = excel_save_path / f"{surname}{name[:1]}-{work_date.month}-{work_date.day}.xlsx"
        wb.save(excel_file_save)
        print(f'Отчёт {excel_file_save} создан')
    else:
        pass
        # print(f'{fio}, не работал {work_date}')
    # вариант с сохранением файла в txt
    # txt_file = excel_save_path / f"{surname}{name[:1]}-{work_date.month}-{work_date.day}.txt"
    # with open(txt_file, 'w') as file:
    #     file.write("'ФИО' | 'ТД,ч' | 'оп.№' | 'T№', 'дата' \n")
    #     for line in queryset:
    #         row = (f"{fio} | {line['norm_calc']} | {line['op_number']} | {line['ws_number']} |"
    #                f"{line['decision_time'].strftime('%d.%m.%Y')} \n")
    #         file.write(row)
    #     # file.write(str(fio_data))
    # pprint(fio_data)
    return fio_data


def file_dirs_clean(dir_path: str) -> list:
    """
    Функция удаляет файлы с расширением extension в директории dir_path относительно главно папки
    :param dir_path: строка относительного пути к файлам
    :return:
    """
    removed = []
    for filename in os.listdir(dir_path):
        file_path = os.path.join(dir_path, filename)
        try:
            if os.path.isfile(file_path):
                os.remove(file_path)
                removed.append(filename)
        except OSError as e:
            print(f"Ошибка удаления файла {file_path}: {e}")
    return removed


def create_fio_report_schedule():
    """
    Формирование отчётов по всем fio из используемой таблицы fio
    для запуска по расписанию
    :return:
    """
    # excel_save_path = Path.joinpath(BASE_DIR, 'daily_reports')
    # print(file_dirs_clean(excel_save_path))
    doers = Doers.objects.all().values('doers')
    today = make_aware(datetime.datetime(year=2024, month=7, day=1))
    # очистка директории в начале месяца
    if today.day == 1:
        excel_save_path = Path.joinpath(BASE_DIR, 'daily_reports')
        print(file_dirs_clean(excel_save_path))
    # today = datetime.datetime.now()
    for fio in doers:
        create_fio_report(fio['doers'], today)
    # return render(request, r"tst/tst.html")


if __name__ == '__main__':
    pass
