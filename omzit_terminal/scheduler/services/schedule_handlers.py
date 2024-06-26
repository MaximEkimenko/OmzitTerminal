import calendar
import json
import shutil
from decimal import Decimal

from django.db.models import Sum
import openpyxl
from PyPDF2 import PdfMerger
from fpdf import FPDF
import datetime
import os

# from omzit_terminal.omzit_terminal.settings import BASE_DIR

from ..models import WorkshopSchedule, ShiftTask
# TODO ФУНКЦИОНАЛ ОЧТЁТОВ ЗАКОНСЕРВИРОВАНО
# from ..models import MonthPlans, DailyReport
import matplotlib.pyplot as plt
import matplotlib
from omzit_terminal.settings import BASE_DIR  # noqa


# TODO подключить logger при расконсервации


def get_done_rate(model_order_query: str) -> float:
    """
    Получение процента готовности заказа
    :param model_order_query: номер заказа
    :return:
    """
    # метод расчёта по количеству принятых сменных заданий
    # TODO реализовать альтернативный метод расчёта по выполненной трудоёмкости.
    total = 0
    all_st = ShiftTask.objects.filter(model_order_query=model_order_query).count()
    done_st = ShiftTask.objects.filter(model_order_query=model_order_query, st_status='принято').count()

    full_td = ShiftTask.objects.filter(
        model_order_query=model_order_query).aggregate(total=Sum('norm_tech'))['total']
    try:
        done_rate = round(100 * (all_st - (all_st - done_st)) / all_st, 2)
    except ZeroDivisionError:
        done_rate = 0

    return done_rate


def get_done_rate_with_td(td: Decimal, model_order: str) -> float:
    """
    Функция расчитывает процент готовности заказ модели по полной трудоёмкости
    :param td: полная трудоёмкость
    :param model_order: заказ модель
    :return:
    """
    if td is None:
        return 0
    done_st = ShiftTask.objects.filter(
        model_order_query=model_order,
        st_status='принято').aggregate(sum=Sum('norm_calc'))['sum']
    if done_st:
        try:
            return round(float(done_st / td) * 100, 1)
        except ZeroDivisionError:
            return 0
    else:
        return 0
    # print(done_st)

    # .aggregate(sum=Sum('norm_calc'))['sum']
    # >> > tt.objects.aggregate(total_likes=Sum('tt_like'))
    # {'total_likes': 0.4470664529184653}


def get_all_done_rate() -> None:
    """
    Получение процента готовности всех заказов, обновление записей done_rate в БД
    :return: None
    """
    all_orders = WorkshopSchedule.objects.values('model_order_query', 'order_status', 'done_rate').distinct()

    for dict_order in all_orders:
        current_order_status = dict_order['order_status']
        current_done_rate = dict_order['done_rate']
        model_order_query = dict_order['model_order_query']
        # TODO оптимизировать! Сделать одним запросом.
        done_rate = get_done_rate(model_order_query)
        # изменение статуса в зависимости от процента готовности при изменении
        if current_order_status != 'завершено' and current_done_rate != done_rate:
            if done_rate == float(0):
                order_status = None
            elif done_rate == float(100):
                order_status = 'завершено'
            else:
                order_status = 'в работе'
            if order_status:
                WorkshopSchedule.objects.filter(model_order_query=model_order_query).update(done_rate=done_rate,
                                                                                            order_status=order_status)


def make_workshop_plan_plot(workshop: int, days_list: list, fact_list: list, aver_fact: float,
                            start_day: datetime, end_day: datetime, yesterday: datetime) -> None:
    """
    Функция для построения инфографики работы цеха
    :param yesterday:
    :param end_day:
    :param start_day:
    :param workshop: цех
    :param days_list: список дней
    :param fact_list: список % фактического выполнения плана цехом по дням
    :param aver_fact: среднее значение факта
    :return: None
    """
    matplotlib.use('agg')  # отключение интерактивного режима
    plt.figure(figsize=(18, 8))  # размер графика
    plt.xlabel('Дата')
    plt.ylabel('% выполнения')
    plt.title(f"График выполнения плана цеха № {workshop} по {yesterday.strftime('%d.%m.%Y')}", fontsize=20)
    plt.yticks(range(0, 300, 10), fontsize=12)  # деления оси y
    fact_color = 'red' if aver_fact < 100 else 'green'  # цвет линии факта
    # корректировка fact_list при небольших значениях
    result_length = 3  # минимальная длина отчёта
    if len(fact_list) < result_length:
        hundred_list = [100] * result_length
        aver_list = [aver_fact] * result_length
        y_axis = [0] * result_length
        x_axis = [(start_day + datetime.timedelta(days=x)).strftime('%d.%m') for x in range(0, end_day.day)]
        for i in range(result_length):
            if i < len(fact_list):
                y_axis[i] = fact_list[i]
            else:
                break
    else:
        y_axis = fact_list
        x_axis = days_list
        result_length = len(fact_list)
        hundred_list = [100] * len(fact_list)
        aver_list = [aver_fact] * len(fact_list)
    plt.tick_params(direction='in', right=True, top=True, left=True)
    plt.plot(x_axis[0:result_length], y_axis, lw=1, color='blue', alpha=1, marker='o', markersize=2)  # факт
    plt.plot(x_axis[0:result_length], aver_list, lw=3, color=fact_color, alpha=1)  # факт средняя
    plt.plot(x_axis[0:result_length], hundred_list, lw=3, color='black', alpha=1)  # 100%
    # вертикальная линия на сегодня
    plt.axvline(x=days_list.index(yesterday.strftime('%d.%m')), color='gray', linestyle='--')
    plt.grid(ls=':')  # сетка графика
    # легенда
    plt.legend(('Кривая выполнения плана', fr'Факт на {days_list[len(days_list) - 1]} - '
                                           fr'{round(aver_fact, 2)}%', '100%'), loc=3, fontsize=16)
    # сохранение графика
    filename = rf'scheduler\static\scheduler\jpg\plan{workshop}-{days_list[0][-2:]}.jpg'
    filename_2 = rf'O:\Расчет эффективности\отчёты цехов\plan{workshop}-{days_list[0][-2:]}.jpg'  # копия в сеть
    filepath = os.path.join(BASE_DIR, filename)
    plt.savefig(fname=filepath, orientation='landscape', bbox_inches='tight', transparent=True)
    # копия в расчёт эффективности
    plt.savefig(fname=filename_2, orientation='landscape', bbox_inches='tight', transparent=True)
    # plt.show()


def create_pdf_report(filename: str, table_data: tuple, image_path: str, header_text: str, footer_text: list,
                      output_dir: str = None, ) -> str:
    """
    Функция создает pdf отчёт о работе цехв из картинки графики и списка табличных данных
    :param footer_text: содержание футера отчёта
    :param header_text: содержание заголовка отчёта
    :param filename: имя файла
    :param table_data: список рядов таблицы, каждый ряд: список ячеек
    :param image_path: путь к картинке графика
    :param output_dir: директория сохранения файла результатов
    :return: полный путь файла
    """

    filepath = os.path.join(output_dir, filename) if output_dir else filename
    # график плана
    pdf = FPDF(orientation='L', unit='mm')
    pdf.add_page()
    pdf.image(image_path, x=0, y=0, w=290)
    pdf.add_font('Arial', '', r'C:\WINDOWS\FONTS\ARIAL.ttf', uni=True)
    pdf.set_font("Arial", '', size=14)
    pdf.ln(130)  # Отступ от графика

    # таблица
    row_height = pdf.font_size + 2  # высота ряда
    spacing = 1  # внутренний отступ таблицы
    col_width = [len(x) * 2 + pdf.font_size * 2.5 + 2 for x in table_data[0]]  # переменная по шапке ширина рядов
    # заголовок
    pdf.cell(sum(col_width), row_height * spacing, txt=header_text, align='C')
    pdf.ln(row_height * spacing)  # отступ заголовка
    for row in table_data:
        for i, item in enumerate(row):
            pdf.cell(col_width[i], row_height * spacing, txt=item, border=1, align='C', ln=0)
        pdf.ln(row_height * spacing)

    # футер
    pdf.add_font('Arial', '', r'C:\WINDOWS\FONTS\ARIAL.ttf', uni=True)
    pdf.set_font("Arial", '', size=10)
    for footer_row in footer_text:
        pdf.ln(row_height * spacing)  # отступ футера
        pdf.cell(sum(col_width), row_height * spacing, txt=footer_row, align='L')

    pdf.output(filepath)
    return filepath


def simple_report_read(exel_file: str) -> dict:
    """
    Функция читает простой файл excel по колонкам дат согласно стандартной форме,
    обновляет данные в общем json,
    читает файл json данными дуги и
    возвращает словарь, который записывается в json
    """
    date = (datetime.datetime.now() - datetime.timedelta(days=1)).date()  # вчера
    try:
        ex_wb = openpyxl.load_workbook(exel_file, data_only=True)  # чтение исходник Excel
    except Exception as e:
        print("Ошибка чтения исходного файла Excel: ", e)
        return dict()
    ex_sh = ex_wb.active
    # определение номера колонки чтения по дате - вчерашний день (число)
    report_day = datetime.datetime.now() - datetime.timedelta(1)
    number_day = int(report_day.strftime('%d'))  # номер дня в формат числа
    day_c1, day_c2, day_c3, day_c4 = 0, 0, 0, 0  # переменные дней
    sum_c1, sum_c2, sum_c3, sum_c4 = 0, 0, 0, 0  # переменные суммы
    for ex_col in ex_sh.iter_cols(min_row=1, max_row=ex_sh.max_row, min_col=1,  # проход по файлу excel
                                  max_col=ex_sh.max_column, values_only=True):
        if ex_col[1] == number_day:
            day_c1 = ex_col[2] if ex_col[2] else 0
            sum_c1 = ex_sh.cell(3, 33).value
            day_c2 = ex_col[3] if ex_col[3] else 0
            sum_c2 = ex_sh.cell(4, 33).value
            day_c3 = ex_col[4] if ex_col[4] else 0
            sum_c3 = ex_sh.cell(5, 33).value
            day_c4 = ex_col[5] if ex_col[5] else 0
            sum_c4 = ex_sh.cell(6, 33).value
    ex_wb.close()
    json_file_path = r'M:\Xranenie\ПТО\1 Екименко М.А'
    json_filename = os.path.join(json_file_path, "all_indicators.json")
    with open(json_filename, 'r') as json_file:
        line = json_file.readline()
        while line:
            if str(date) in line:
                json_file_content = json.loads(line)
                break
            line = json_file.readline()
    try:
        json_file_content
    except Exception as e:
        print("Ошибка чтения файла json: ", e)
        return dict()
    # обновление
    if "Брак" in exel_file:
        json_file_content[str(date)].update(
            [('c1_fails_day', (day_c1, sum_c1)), ('c1_fails_result', f'{day_c1} / {sum_c1}'),
             ('c2_fails_day', (day_c2, sum_c2)), ('c2_fails_result', f'{day_c2} / {sum_c2}'),
             ('c3_fails_day', (day_c3, sum_c3)), ('c3_fails_result', f'{day_c3} / {sum_c3}'),
             ('c4_fails_day', (day_c4, sum_c4)), ('c4_fails_result', f'{day_c4} / {sum_c4}'),
             ])
    if "ТБ" in exel_file:
        json_file_content[str(date)].update(
            [('c1_un_save_day', (day_c1, sum_c1)), ('c1_un_save_result', f'{day_c1} / {sum_c1}'),
             ('c2_un_save_day', (day_c2, sum_c2)), ('c2_un_save_result', f'{day_c2} / {sum_c2}'),
             ('c3_un_save_day', (day_c3, sum_c3)), ('c3_un_save_result', f'{day_c3} / {sum_c3}'),
             ('c4_un_save_day', (day_c4, sum_c4)), ('c4_un_save_result', f'{day_c4} / {sum_c4}'),
             ])
    # обновление json
    with open(json_filename, 'w') as json_file:
        json.dump(json_file_content, json_file)
    return json_file_content


def report_merger(month: int, merge_dir: str = None) -> None:
    """
    Функция принимает месяц в который нужно объединить отчёты и директорию, где хранятся отчёты.
    Результирующий файл сохраняется в ту же директорию
    :param month:
    :param merge_dir:
    :return: None
    """

    x = [os.path.join(merge_dir, file) for file in os.listdir(merge_dir) if file.endswith(f"{month}.pdf")]
    merger = PdfMerger()
    for pdf in x:
        merger.append(open(pdf, 'rb'))
    with open(os.path.join(merge_dir, f"Отчёт за {month} месяц.pdf"), "wb") as result_file:
        merger.write(result_file)


def report_merger_schedule() -> None:
    """
    Функция объединения отчётов для запуска по расписанию
    :return:
    """
    month = (datetime.datetime.now() - datetime.timedelta(days=1)).date().month  # месяц вчера
    pdf_path = os.path.join(BASE_DIR, r'scheduler\static\scheduler\pdf')
    try:
        report_merger(merge_dir=pdf_path, month=month)
        print('Отчёты объеденины.')
    except Exception as e:
        print('Ошибка объединения отчётов: ', e)
    # копия в расчёт эффективности
    shutil.copy(os.path.join(pdf_path, f"Отчёт за {month} месяц.pdf"), os.path.join(r"O:\Расчет эффективности\Рапорты",
                                                                                    f"Отчёт за {month} месяц.pdf"))


def report_json_create_schedule():
    """
    Функция обновления в данных ОТПБ в отчёте для запуска по расписанию
    :return:
    """
    otk_excel_file = r'M:\Xranenie\ПТО\1 Екименко М.А\Брак.xlsx'
    otpb_excel_file = r'M:\Xranenie\ПТО\1 Екименко М.А\Нарушения ТБ.xlsx'
    try:
        simple_report_read(otk_excel_file)
        print('json otk_excel_file info updated')
    except Exception as e:
        print(e, ' json update failed at otk_excel_file read!')
    try:
        simple_report_read(otpb_excel_file)
        print('json otpb_excel_file info updated')
    except Exception as e:
        print(e, ' json update failed at otpb_excel_file read!')

# TODO ФУНКЦИОНАЛ ОЧТЁТОВ ЗАКОНСЕРВИРОВАНО
# def days_report_create() -> None:
#     """
#     Функция создает записи по каждому дню в начале месяца. Функция запускается по расписанию первого
#      числа каждого месяца.
#     :return: None
#     """
#     start_date = datetime.datetime.now().replace(day=1)  # первый день текущего месяца
#     _, last_day = calendar.monthrange(start_date.year, start_date.month)  # последний день текущего месяца
#     end_date = datetime.datetime(year=start_date.year, month=start_date.month, day=last_day)  # последняя дата месяца)
#     # список дат дней месяца
#     daterange = [(start_date.date() + datetime.timedelta(days=x)) for x in range(0, end_date.day)]
#     days_dict = dict()
#     days_list = []
#     for i in range(1, 4 + 1):  # цикл по всем цехам
#         # создание записей планов месяца
#         try:
#             if not MonthPlans.objects.get(workshop=i, month_plan=daterange[0]):
#                 MonthPlans.objects.create(workshop=i, month_plan=daterange[0])
#         except Exception:
#             MonthPlans.objects.create(workshop=i, month_plan=daterange[0])
#         month_plan_data = MonthPlans.objects.get(workshop=i, month_plan=daterange[0])  # план цеха
#         for date in daterange:
#             days_dict.update({'calendar_day': date, 'workshop': i, 'month_plan_data': month_plan_data})
#             days_list.append(DailyReport(**days_dict))
#     # создание записей каждого дня месяца
#     try:
#         if not DailyReport.objects.get(calendar_day=daterange[0], workshop=1):
#             DailyReport.objects.bulk_create(days_list)
#     except Exception:
#         DailyReport.objects.bulk_create(days_list)
