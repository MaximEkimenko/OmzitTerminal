import datetime
import os
import shutil
from typing import Tuple

import openpyxl
import pandas
from openpyxl.styles import Font
from django.core.mail import EmailMessage
from django.db.models import Sum
from django.utils.timezone import make_aware, make_naive

from omzit_terminal.settings import BASE_DIR
from scheduler.models import ShiftTask, Doers

valentina_id = 6011624527  # тест
shkapov_id = 1890988322  # Шкапов
omzit_chat_id = -1001507361668  # группы
omzit_otk_chat_id = -981440150
omzit_bot_id_report = 6015644097  # боты
omzit_bot_id_otk = 6337176450
savchenko_id = 2131171377
korneychuk_id = 1014915479
posohov_id = 2051721470  # цех 1
gordii_id = 6374431046
ermishkin_id = 5221029965
kondratiev_id = 6125791135
achmetov_id = 1153114403
kutorov_id = 5382773351  # цех 2
mailashov_id = 546976234
gorojanski_id = 6299557037
kulbashin_id = 5426476877
pospelov_id = 1377896858
skorobogatov_id = 5439414299
rihmaer_id = 6305730497
ostrijnoi_id = 5380143506
fursov_id = 5783679882  # цех 4
sergeev_id = 5209291574
loshkov_id = 6243031150
eihgorn_id = 5780975373
lipski_id = 6424114889
lubimov_id = 5705407706
donskaya_id = 6359131276  # ОТК
averkina_id = 1563020113
dolganev_id = 1907891961
potapova_id = 5010645397
sofinskaya_id = 1358370501
dubenuk_id = 1359982302
shaparenko_id = 5283718961
galai_id = 6591032501
shagov_id = 1906275223
tashbulatov_id = 6072981051
vafin_id = 1419051027
sheglov_id = 1501419738
kalashnikov_id = 1121811565
sultigova_id = 6049253475
voronin_id = 6247745541  # плазма
makeev_id = 258500986
godenchuk_id = 5400346808
erin_id = 1293843639  # цех 3
kozukin_id = 1148106959
procenko_id = 5782500917
hasanov_id = 6829564468

id_fios = {
    valentina_id: 'Фадеева В.С.',  # Тест
    procenko_id: 'Проценко В.Б.',
    shkapov_id: 'Шкапов Д.А.',
    savchenko_id: 'Савченко Е.Н.',
    donskaya_id: 'Донская Ю.Г.',
    averkina_id: 'Аверкина О.В.',  # ОТК
    dolganev_id: 'Долганев А.Н.',
    potapova_id: 'Потапова М. А.',
    sofinskaya_id: 'Софинская А. Г.',
    dubenuk_id: 'Дубенюк А. П.',
    shaparenko_id: 'Шапаренко Т.',
    galai_id: 'Гайлай В.',
    shagov_id: 'Шагов И.',
    tashbulatov_id: 'Ташбулатов Н.',
    vafin_id: 'Вафин Р.',
    sheglov_id: 'Щеглов В.',
    kalashnikov_id: 'Калашников Д.',
    sultigova_id: 'Султыгова О.',
    posohov_id: 'Посохов О.С.',  # цех 1
    gordii_id: 'Гордий В.В.',
    ermishkin_id: 'Ермишкин В.М.',
    kondratiev_id: 'Кондратьев П.В.',
    achmetov_id: 'Ахметов К.',
    kutorov_id: 'Куторов В.В.',  # цех 2
    mailashov_id: 'Майлашов О.',
    gorojanski_id: 'Горожанский Н.Н.',
    pospelov_id: 'Поспелов К.С.',
    kulbashin_id: 'Кульбашин Ю.А.',
    skorobogatov_id: 'Скоробогатов А.',
    ostrijnoi_id: 'Острижной К.',
    rihmaer_id: 'Рихмаер Ю.С.',
    erin_id: 'Ерин К.В.',  # цех 3
    kozukin_id: 'Козюкин М.В.',
    hasanov_id: 'Хасанов Е.Х.',
    korneychuk_id: 'Корнейчук Д.А.',  # цех 4
    fursov_id: 'Фурсов А.П.',
    sergeev_id: 'Сергеев В.Б.',
    loshkov_id: 'Лошков А.В.',
    eihgorn_id: 'Эйхгорн Р.В.',
    lipski_id: 'Липский В.',
    lubimov_id: 'Любимов Д.',
    voronin_id: 'Воронин И.',  # плазма
    makeev_id: 'Макеев И.',
    godenchuk_id: 'Годенчук А.Л.'
}


def shift_tasks_auto_report():  # TODO перенести в service
    """
    Отправляет отчет по сменным заданиям на электронную почту и в папку O:/Расчет эффективности/Отчёты по СЗ
    """
    start = make_aware(datetime.datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0))
    end = make_aware(datetime.datetime.now())
    exel_file = create_shift_task_report(start, end)
    shutil.copy(exel_file, os.path.join(r"O:\Расчет эффективности\Отчёты по СЗ", os.path.basename(exel_file)))
    email = EmailMessage(
        f"Отчет {start.strftime('%d.%m.%Y')}-{end.strftime('%d.%m.%Y')}",
        f"Отчет {start.strftime('%d.%m.%Y')}-{end.strftime('%d.%m.%Y')}",
        "omzit-report@yandex.ru",
        [
            "alex4ekalovets@gmail.com",
            "pdo02@omzit.ru",
            "pdo06@omzit.ru",
            "pdo09@omzit.ru",
            "e.savchenko@omzit.ru",
            "PVB@omzit.ru",
            "m.ekimenko@omzit.ru"
        ],
        [],
    )
    email.attach_file(exel_file)
    email.send()


def create_shift_task_report(start: datetime, end: datetime) -> str:  # TODO перенести в service
    """
    Создает excel-файл отчета по сменным заданиям в папке xslx в корне проекта
    :param start: с даты (дата распределения)
    :param end: по дату (дата распределения)
    :return:
    """
    # Формируем имена столбцов для полного отчета из аттрибута модели verbose_name

    verbose_names = dict()
    for field in ShiftTask._meta.get_fields():
        if hasattr(field, "verbose_name"):
            verbose_names[field.name] = field.verbose_name
        else:
            verbose_names[field.name] = field.name

    queryset = ShiftTask.objects.exclude(
        fio_doer="не распределено"
    ).order_by("datetime_assign_wp")

    fields_1C_report = (
        "pk",  # №
        "model_name",  # модель
        "order",  # заказ
        "op_number",  # № Операции
        "op_name_full",  # Операция
        "fio_doer",  # Исполнители
        "decision_time",  # Дата готовности
        "st_status",  # статус СЗ
    )
    fields_disp_report = (
        "pk",  # №
        "model_name",  # модель
        "order",  # заказ
        "ws_number",  # РЦ
        "op_number",  # № Операции
        "op_name_full",  # Операция
        "fio_doer",  # Исполнители
        "datetime_assign_wp",  # Дата распределения
        "datetime_job_start",  # Дата начала
        "decision_time",  # Дата окончания
        "job_duration",  # Длительность работы
        "norm_tech",  # Технологическая норма
        "doers_tech",  # Количество исполнителей по ТД
        "norm_calc",  # Расчетная норма
        "st_status",  # Статус СЗ
        "master_finish_wp",  # Мастер
        "otk_decision",  # Контролер
    )

    # Определяем путь к excel файлу шаблона
    exel_file_src = BASE_DIR / "ReportTemplate.xlsx"
    # Формируем название нового файла
    new_file_name = (f"{datetime.datetime.now().strftime('%Y.%m.%d')} report "
                     f"{start.strftime('%d.%m.%Y')}-{end.strftime('%d.%m.%Y')}.xlsx")
    # Создаем папку для хранения отчетов
    if not os.path.exists(BASE_DIR / "xlsx"):
        os.mkdir(BASE_DIR / "xlsx")
    # Формируем путь к новому файлу
    exel_file_dst = BASE_DIR / "xlsx" / new_file_name
    # Копируем шаблон в новый файл отчета
    shutil.copy(exel_file_src, exel_file_dst)

    # Формируем отчет
    ex_wb = openpyxl.load_workbook(exel_file_src, data_only=True)
    sheets_reports = {
        "Отчет для 1С": queryset.values(*fields_1C_report).filter(
            datetime_assign_wp__gte=start,
            datetime_assign_wp__lte=end
        ),
        "Отчет для диспетчера": queryset.values(*fields_disp_report).filter(
            datetime_assign_wp__gte=start,
            datetime_assign_wp__lte=end
        ),
        "Полный отчет": queryset.values(*verbose_names),
        "Отчет по выполненной норме": fio_st_time_counter(start, end),
    }
    for sheet_name in sheets_reports:
        ex_sh = ex_wb[sheet_name]
        report = sheets_reports[sheet_name]
        font = Font(name='Arial', size=12)
        if report:
            # Для полного отчета создаем шапку из verbose_name
            if sheet_name == "Полный отчет":
                for i, key in enumerate(report[0]):
                    ex_sh.cell(row=1, column=i + 1).value = verbose_names[key]
            # Заполняем строки данными
            for i, row in enumerate(report):
                for j, key in enumerate(row):
                    if key in ["master_finish_wp", "otk_decision"] and row[key] and row[key].isdigit():
                        row[key] = id_fios.get(int(row[key]), int(row[key]))
                    cell = ex_sh.cell(row=i + 2, column=j + 1)
                    try:
                        row[key] = make_naive(row[key])
                    except Exception:
                        pass
                    cell.value = row[key]
                    cell.font = font
            ex_wb.save(exel_file_dst)
    return exel_file_dst


def fio_st_time_counter(start: datetime, end: datetime):
    """
    Отчет по количеству часов по сменным заданиям по исполнителям
    """
    months = ['', 'Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь', 'Июль', 'Август', 'Сентябрь', 'Октябрь',
              'Ноябрь', 'Декабрь']
    month = months[int(start.strftime('%m'))]
    year = int(start.strftime('%Y'))
    doer_job_duration = []
    doers = Doers.objects.values_list('doers', flat=True).order_by('doers').all()
    shift_tasks = ShiftTask.objects.filter(decision_time__gte=start, decision_time__lte=end)
    main_path = fr"M:\Xranenie\Расчет эффективности"
    cex_1_timesheets = os.path.join(main_path, 'Табель цеха №1.xlsx')
    cex_2_timesheets = os.path.join(main_path, 'Табель цеха №2.xlsx')
    cex_1 = None
    cex_2 = None
    try:
        cex_1 = pandas.read_excel(cex_1_timesheets, sheet_name=f'{month} {year}')
        cex_2 = pandas.read_excel(cex_2_timesheets, sheet_name=f'{month} {year}')
    except Exception as ex:
        print(ex)
    for i, doer in enumerate(doers):
        sum_job_duration = shift_tasks.filter(
            fio_doer__contains=doer, st_status='принято'
        ).aggregate(duration=Sum('norm_calc'))
        data = {
            "fio": doer,
            "duration": sum_job_duration['duration']
        }
        if cex_1 is not None:
            try:
                data["cex1"] = int(cex_1[cex_1.iloc[:, 1] == doer].iloc[:, 37].iloc[0])
            except Exception as ex:
                print(ex)
                data["cex1"] = 'Нет в табеле'
        else:
            data["cex1"] = f'Файл {cex_1_timesheets} или вкладка {month} {year} недоступны'
        if cex_2 is not None:
            try:
                data["cex2"] = int(cex_2[cex_2.iloc[:, 1] == doer].iloc[:, 37].iloc[0])
            except Exception as ex:
                print(ex)
                data["cex2"] = 'Нет в табеле'
        else:
            data["cex2"] = f'Файл {cex_2_timesheets} или вкладка {month} {year} недоступны'
        doer_job_duration.append(data)
    return doer_job_duration


def get_start_end_st_report(start: str, end: str) -> Tuple:  # TODO перенести в service
    """
    Преобразует полученные от пользователя строки с датами или null в дату и время
    :param start: с даты (Дата распределения)
    :param end: по дату (Дата распределения)
    :return: дату начала, дату окончания формирования отчета
    """
    if start == "null":
        start = make_aware(datetime.datetime(year=1990, month=1, day=1, hour=0, minute=0, second=0, microsecond=0))
    else:
        start = make_aware(datetime.datetime.strptime(start, "%d.%m.%Y"))
    if end == "null":
        end = make_aware(datetime.datetime.now())
    else:
        end = make_aware(datetime.datetime.strptime(end, "%d.%m.%Y").replace(hour=23, minute=59, second=59))
    return start, end
