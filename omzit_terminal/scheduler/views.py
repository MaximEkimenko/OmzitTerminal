# import calendar
# import json
import asyncio
import datetime
import json
import os
import shutil
import threading
from typing import Tuple
# from collections import OrderedDict

import openpyxl
import psutil
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.views import LoginView
from django.contrib.auth.forms import AuthenticationForm
from django.core.exceptions import PermissionDenied
from django.core.mail import EmailMessage
from django.db import IntegrityError
from django.http import FileResponse, JsonResponse

from django.shortcuts import render, redirect
from django.urls import reverse_lazy

from django.db.models import Q, Avg, Sum
from django.utils.timezone import make_aware, make_naive

from omzit_terminal.settings import BASE_DIR
from tehnolog.services.service_handlers import handle_uploaded_file
from .filters import get_filterset
from .forms import (SchedulerWorkshop, SchedulerWorkplace, FioDoer, QueryDraw, PlanBid, ReportForm,
                    CdwChoiceForm, SendSZForm)
# from .forms import DailyReportForm, PlanResortHiddenForm # TODO функционал отчётов ЗАКОНСЕРВИРОВАНО
# from .models import  DailyReport, MonthPlans # TODO функционал отчётов ЗАКОНСЕРВИРОВАНО
from .models import WorkshopSchedule, ShiftTask

from .services.schedule_handlers import get_all_done_rate, make_workshop_plan_plot, create_pdf_report, report_merger
from worker.services.master_call_function import terminal_message_to_id
from .services.specification import connect_to_client, get_specifications_server, get_specifications_ssh
from .services.sz_to_pdf import create_pdf_sz

SPEC_CREATION_PROCESS = dict()
SPEC_CREATION_THREAD = dict()
TERMINAL_GROUP_ID = os.getenv('TERMINAL_GROUP_ID')


@login_required(login_url="login")
def scheduler(request):
    """
        Планирование графика цеха и создание запросов на КД
        :param request:
        :return:
        """
    if str(request.user.username).strip()[:5] != "admin" and str(request.user.username[:4]).strip() != "disp":
        raise PermissionDenied
    group_id = TERMINAL_GROUP_ID  # тг группа

    # обновление процента готовности всех заказов
    # TODO модифицировать расчёт процента готовности всех заказов по взвешенной трудоёмкости
    #  сделать невозможным заполнять запрос с кириллицей
    get_all_done_rate()
    # график изделий
    workshop_schedule_fields = ('workshop', 'order', 'model_name', 'datetime_done', 'order_status', 'done_rate')
    workshop_schedule = (WorkshopSchedule.objects.values(*workshop_schedule_fields)
                         .exclude(datetime_done=None).exclude(order_status='не запланировано')
                         .order_by('datetime_done'))
    # фильтры в колонки графика
    f_w = get_filterset(data=request.GET, queryset=workshop_schedule, fields=workshop_schedule_fields, index=1)
    # перечень запросов на КД
    td_queries_fields = ('model_order_query', 'query_prior', 'td_status', 'order_status')  # поля таблицы
    td_queries = (WorkshopSchedule.objects.values(*td_queries_fields).exclude(td_status='завершено'))
    # фильтры в колонки заявок
    # f_q = get_filterset_second_table(data=request.GET, queryset=td_queries, fields=td_queries_fields)
    f_q = get_filterset(data=request.GET, queryset=td_queries, fields=td_queries_fields, index=2)

    # форма запроса КД
    form_query_draw = QueryDraw()
    if request.method == 'POST':
        form_workshop_plan = SchedulerWorkshop(request.POST)
        if form_workshop_plan.is_valid():
            print(form_workshop_plan.cleaned_data)
            # заполнение графика цеха датой готовности и цехом
            try:
                # Планирование графика цеха
                # заполнение срока готовности, категория изделия, статус изделия, ФИО планировщика
                (WorkshopSchedule.objects.filter(model_order_query=form_workshop_plan.
                                                 cleaned_data['model_order_query'].model_order_query).update(
                    datetime_done=form_workshop_plan.cleaned_data['datetime_done'],
                    workshop=form_workshop_plan.cleaned_data['workshop'],
                    product_category=str(form_workshop_plan.cleaned_data['category']),
                    order_status='запланировано',
                    dispatcher_plan_ws_fio=f'{request.user.first_name} {request.user.last_name}'))
                # Заполнение данных СЗ, статус СЗ, ФИО планировщика, категория изделия,
                alert = 'Данные в график успешно занесены! '
                print('Данные в график успешно занесены!\n')
                # заполнение модели ShiftTask данными планирования цехов
                print(ShiftTask.objects.filter(model_order_query=form_workshop_plan.cleaned_data['model_order_query']))
                (ShiftTask.objects.filter(
                    model_order_query=form_workshop_plan.cleaned_data['model_order_query'].model_order_query)
                 .update(workshop=form_workshop_plan.cleaned_data['workshop'],
                         datetime_done=form_workshop_plan.cleaned_data['datetime_done'],
                         product_category=str(form_workshop_plan.cleaned_data['category']),
                         ))
                print('Данные сменного задания успешно занесены!')
                alert += 'Данные сменного задания успешно занесены.'
                context = {'form_workshop_plan': form_workshop_plan, 'td_queries': td_queries, 'alert': alert,
                           'workshop_schedule': workshop_schedule, 'form_query_draw': form_query_draw,
                           'filter_w': f_w, 'filter_q': f_q}
                # сообщение в группу
                success_group_message = (f"Заказ-модель: "
                                         f"{form_workshop_plan.cleaned_data['model_order_query'].model_order_query} "
                                         f"успешно запланирован на {form_workshop_plan.cleaned_data['datetime_done']}. "
                                         f"Запланировал: {request.user.first_name} {request.user.last_name}.")
                asyncio.run(terminal_message_to_id(to_id=group_id, text_message_to_id=success_group_message))
            except Exception as e:
                print(e, ' Ошибка запаси в базу SchedulerWorkshop')
                alert = f'Ошибка занесения данных.'
                context = {'form_workshop_plan': form_workshop_plan, 'workshop_schedule': workshop_schedule,
                           'td_queries': td_queries, 'form_query_draw': form_query_draw, 'alert': alert,
                           'filter_w': f_w, 'filter_q': f_q}
                return render(request, r"scheduler/scheduler.html", context=context)
        else:
            context = {'form_workshop_plan': form_workshop_plan, 'workshop_schedule': workshop_schedule,
                       'td_queries': td_queries, 'form_query_draw': form_query_draw,
                       'filter_w': f_w, 'filter_q': f_q}
    else:
        # чистые формы для первого запуска
        form_workshop_plan = SchedulerWorkshop()
        context = {'form_workshop_plan': form_workshop_plan, 'workshop_schedule': workshop_schedule,
                   'td_queries': td_queries, 'form_query_draw': form_query_draw,
                   'filter_w': f_w, 'filter_q': f_q}
    return render(request, r"scheduler/scheduler.html", context=context)


@login_required(login_url="login")
def td_query(request):
    """
    Обработка формы запросы чертежей
    :param request:
    :return:
    """
    group_id = TERMINAL_GROUP_ID  # тг группа
    if request.method == 'POST':
        form_query_draw = QueryDraw(request.POST)
        if form_query_draw.is_valid():
            # модель_заказ
            model_order_query = (f"{form_query_draw.cleaned_data['order_query'].strip()}_"
                                 f"{form_query_draw.cleaned_data['model_query'].strip()}")
            # заполнение графика цеха данными запроса на чертеж
            if not (WorkshopSchedule.objects.filter(order=form_query_draw.cleaned_data['order_query'].strip(),
                                                    model_order_query=model_order_query,
                                                    query_prior=form_query_draw.cleaned_data
                                                    ['query_prior'],
                                                    model_name=form_query_draw.cleaned_data
                                                    ['model_query'].strip()).exists()):
                WorkshopSchedule.objects.create(order=form_query_draw.cleaned_data['order_query'].strip(),
                                                model_order_query=model_order_query,
                                                query_prior=form_query_draw.cleaned_data
                                                ['query_prior'],
                                                model_name=form_query_draw.cleaned_data
                                                ['model_query'].strip(),
                                                td_status="запрошено",
                                                dispatcher_query_td_fio=f"{request.user.first_name} "
                                                                        f"{request.user.last_name}")

                # сообщение об успехе для отправки в группу
                success_group_message = (f"Поступила заявка на КД. Изделие: "
                                         f"{form_query_draw.cleaned_data['model_query']}. Заказ:  "
                                         f"{form_query_draw.cleaned_data['order_query']}. Приоритет: "
                                         f"{form_query_draw.cleaned_data['query_prior']}. "
                                         f"Заявку составил: {request.user.first_name} {request.user.last_name}.")
                # asyncio.run(terminal_message_to_id(to_id=group_id, text_message_to_id=success_group_message))
                # создание папки в общем доступе для чертежей модели
                if not os.path.exists(rf'C:\draws\{model_order_query}'):
                    os.mkdir(rf'C:\draws\{model_order_query}')
        else:
            pass

    return redirect('scheduler')  # обновление страницы при успехе


@login_required(login_url="login")
def schedulerwp(request):
    """
    Планирование графика РЦ
    :param request:
    :return:
    """
    # отображение графика РЦ
    # выборка из уже занесенного
    if str(request.user.username).strip()[:5] != "admin" and str(request.user.username[:4]).strip() != "disp":
        raise PermissionDenied
    shift_task_fields = (
        'id',
        'workshop',
        'order',
        'model_name',
        'datetime_done',
        'ws_number',
        'op_number',
        'op_name_full',
        'norm_tech',
        'fio_doer',
        'st_status'
    )
    workplace_schedule = (ShiftTask.objects.values(*shift_task_fields).all().exclude(datetime_done=None)
                          .order_by("ws_number", "model_name", ))
    alert_message = ''
    if request.method == 'POST':
        form_workplace_plan = SchedulerWorkplace(request.POST)
        form_report = ReportForm()
        if form_workplace_plan.is_valid():
            print(form_workplace_plan.cleaned_data)
            ws_number = form_workplace_plan.cleaned_data['ws_number'].ws_number
            model_order_query = form_workplace_plan.cleaned_data['model_order_query'].model_order_query
            try:
                filtered_workplace_schedule = (
                    ShiftTask.objects.values(*shift_task_fields)
                    .filter(ws_number=str(ws_number), model_order_query=model_order_query, next_shift_task=None)
                    .filter(Q(fio_doer='не распределено') | Q(st_status='брак') | Q(st_status='не принято'))
                )
            except Exception as e:
                filtered_workplace_schedule = dict()
                print('Ошибка получения filtered_workplace_schedule', e)
            # if filtered_workplace_schedule:
            return redirect(f'/scheduler/schedulerfio{ws_number}_{model_order_query}')
            # else:
            #     alert_message = f'Для Т{ws_number} на {model_order_query} нераспределённые задания отсутствуют.'
            #     form_workplace_plan = SchedulerWorkplace()
    else:
        form_workplace_plan = SchedulerWorkplace()
        form_report = ReportForm(initial={'date_end': datetime.datetime.now(),
                                          'date_start': datetime.datetime.now()
                                 .replace(day=1, hour=0, minute=0, second=0, microsecond=0)})
    context = {
        'workplace_schedule': workplace_schedule,
        'form_workplace_plan': form_workplace_plan,
        'form_report': form_report,
        'alert_message': alert_message,
    }
    return render(request, fr"schedulerwp/schedulerwp.html", context=context)


@login_required(login_url="login")
def schedulerfio(request, ws_number, model_order_query):
    """
    Распределение ФИО на РЦ
    :param model_order_query:
    :param ws_number:
    :param request:
    :return:
    """
    if str(request.user.username).strip()[:5] != "admin" and str(request.user.username[:4]).strip() != "disp":
        raise PermissionDenied

    shift_task_fields = (
        'id', 'workshop', 'order', 'model_name', 'datetime_done', 'ws_number', 'op_number', 'op_name_full',
        'norm_tech', 'fio_doer', 'st_status', 'plasma_layout'
    )
    # определения рабочего центра и id
    if not request.user.username:  # если не авторизован, то отправляется на авторизацию
        return redirect('login/')
    try:
        filtered_workplace_schedule = (
            ShiftTask.objects.values(
                *shift_task_fields, 'workpiece__text', 'workpiece__layouts_done', 'workpiece__count'
            )
            .exclude(st_status='раскладка')
            .filter(ws_number=str(ws_number), model_order_query=model_order_query, next_shift_task=None)
            .filter(Q(fio_doer='не распределено') | Q(st_status='брак') | Q(st_status='не принято'))
        )
        f = get_filterset(data=request.GET, queryset=filtered_workplace_schedule, fields=shift_task_fields)
    except Exception as e:
        filtered_workplace_schedule = dict()
        print('Ошибка получения filtered_workplace_schedule', e)
    success = 1
    alert_message = ''
    action = None  # действие по нажатию кнопки в POST форме
    layout = None  # номер раскладки
    pk = None  # id сменного задания
    percentages = None
    fios_doers = None

    form_fio_doer = FioDoer()

    if request.method == 'POST':
        form_submit = request.POST.get("form", "")  # форма по которой выполнен submit
        if "change" in form_submit:
            filtered_workplace_schedule = (
                ShiftTask.objects.values(*shift_task_fields, 'workpiece__text', 'workpiece__layouts_done',
                                         'workpiece__count')
                .exclude(st_status='раскладка').exclude(fio_doer='не распределено')
                .filter(ws_number=str(ws_number), model_order_query=model_order_query, next_shift_task=None)
            )
            action = 'change_distribution'

        elif 'confirm' in form_submit:
            _, pk, layout = form_submit.split("|")
            form_fio_doer = FioDoer(request.POST)
            if form_fio_doer.is_valid():
                # Получение списка без None
                fios = list(filter(
                    lambda x: x != 'None',
                    (str(form_fio_doer.cleaned_data[f'fio_{i}']) for i in range(1, 5))
                ))
                unique_fios = set(fios)
                doers_fios = ', '.join(unique_fios)  # получение уникального списка
                print('DOERS-', doers_fios)
                if len(fios) == len(unique_fios):  # если нет повторений в списке fios
                    if 'redistribute' in form_submit:
                        data = {
                            'fio_doer': doers_fios,
                            'master_assign_wp_fio': f'{request.user.first_name}'
                        }
                    else:
                        data = {
                            'fio_doer': doers_fios,
                            'datetime_assign_wp': make_aware(datetime.datetime.now()),
                            'st_status': 'запланировано',
                            'datetime_job_start': None,
                            'decision_time': None,
                            'master_assign_wp_fio': f'{request.user.first_name}'
                        }
                    if layout != '':  # если распределяем по номеру раскладки
                        # находим все сменные задания, где раскладка является ключом в выполненных раскладках
                        if 'redistribute' in form_submit:
                            shift_tasks = ShiftTask.objects.filter(
                                plasma_layout=layout).exclude(fio_doer='не распределено')
                            for shift_task in shift_tasks:
                                shift_task.workpiece["fio_percentages"] = [
                                    form_fio_doer.cleaned_data[f'fio_{i}_percentage'] for i in range(1, 5)
                                ]
                                shift_task.fio_doer = doers_fios
                                shift_task.save()
                        else:
                            shift_tasks = ShiftTask.objects.filter(
                                workpiece__layouts_done__icontains=f'"{layout}":',
                                fio_doer='не распределено',
                            )
                            for shift_task in shift_tasks:
                                workpiece = shift_task.workpiece
                                workpiece["fio_percentages"] = [
                                    form_fio_doer.cleaned_data[f'fio_{i}_percentage'] for i in range(1, 5)
                                ]
                                # если раскладка на деталь одна или последняя и полностью закрывает потребность
                                # в количестве детали или в раскладке больше, то назначаем исполнителей
                                # на текущее сменное задание
                                all_layouts_done = len(workpiece['layouts']) == 0 and len(
                                    workpiece['layouts_done']) == 1
                                is_enough = int(workpiece['layouts_total']) >= int(workpiece['count'])
                                if all_layouts_done and is_enough:
                                    data['norm_tech'] = workpiece['layouts_done'][layout]['total_time']
                                    workpiece.update({
                                        'layouts': {},
                                        'layouts_done': {},
                                    })
                                    data['workpiece'] = workpiece
                                    data['plasma_layout'] = layout
                                    for field, value in data.items():
                                        setattr(shift_task, field, value)
                                    shift_task.save()
                                else:
                                    layout_data = workpiece['layouts_done'].pop(layout)
                                    layout_count = sum(layout_data['count'])
                                    workpiece['layouts_total'] -= layout_count
                                    workpiece['count'] -= layout_count
                                    shift_task.workpiece = workpiece
                                    if len(workpiece['layouts_done']) == 0:
                                        shift_task.st_status = 'раскладка'
                                    shift_task.save()

                                    new_shift_task = ShiftTask.objects.get(pk=shift_task.id)
                                    new_shift_task.pk = None

                                    workpiece.update({
                                        'count': layout_count,
                                        'layouts': {},
                                        'layouts_done': {},
                                        'layouts_total': layout_count
                                    })
                                    data['workpiece'] = workpiece
                                    data['plasma_layout'] = layout
                                    data['norm_tech'] = layout_data['total_time']
                                    for field, value in data.items():
                                        setattr(new_shift_task, field, value)
                                    new_shift_task.save()
                    elif pk != '':  # если распределяем по id сменного задания
                        shift_task = ShiftTask.objects.get(pk=int(pk))
                        if shift_task.st_status == "брак":
                            #  создаем дубликат СЗ с браком
                            new_shift_task = ShiftTask.objects.get(pk=int(pk))
                            new_shift_task.pk = None
                            for field, value in data.items():
                                setattr(new_shift_task, field, value)
                            new_shift_task.save()
                            #  добавляем в СЗ с браком ссылку на новое СЗ для исправления брака
                            shift_task.next_shift_task = new_shift_task
                        else:  # первичное распределение СЗ
                            for field, value in data.items():
                                setattr(shift_task, field, value)
                        fio_percentages = {
                            "fio_percentages": [form_fio_doer.cleaned_data[f'fio_{i}_percentage'] for i in range(1, 5)],
                        }
                        if shift_task.workpiece:
                            shift_task.workpiece.update(fio_percentages)
                        else:
                            shift_task.workpiece = fio_percentages
                        shift_task.save()

                    alert_message = f'Успешно распределено!'
                else:  # если есть повторения в списке fios
                    alert_message = f'Исполнители дублируются. Измените исполнителей.'
                    success = 0
            pk = layout = None

        elif "redistribute" in form_submit:
            filtered_workplace_schedule = (
                ShiftTask.objects.values(*shift_task_fields, 'workpiece__text', 'workpiece__layouts_done',
                                         'workpiece__count')
                .exclude(st_status='раскладка').exclude(fio_doer='не распределено')
                .filter(ws_number=str(ws_number), model_order_query=model_order_query, next_shift_task=None)
            )
            _, pk, layout = form_submit.split("|")
            shift_tasks = ShiftTask.objects.values_list('workpiece__fio_percentages', 'fio_doer')
            if layout != '':
                percentages, fios = shift_tasks.filter(plasma_layout=layout)[0]
                filtered_workplace_schedule = filtered_workplace_schedule.filter(plasma_layout=layout)
            else:
                percentages, fios = shift_tasks.filter(pk=pk)[0]
                filtered_workplace_schedule = filtered_workplace_schedule.filter(pk=pk)
            fios_doers = fios.split(', ')
            action = 'redistribute'

        elif 'distribute' in form_submit:
            _, pk, layout = form_submit.split("|")
            if layout != '':
                filtered_workplace_schedule = filtered_workplace_schedule.filter(
                    workpiece__layouts_done__icontains=f'"{layout}":'
                )
            else:
                filtered_workplace_schedule = filtered_workplace_schedule.filter(pk=pk)
            action = 'distribute'

        f = get_filterset(data=request.GET, queryset=filtered_workplace_schedule, fields=shift_task_fields)

    context = {
        'filtered_workplace_schedule': filtered_workplace_schedule,
        'form_fio_doer': form_fio_doer,
        'alert_message': alert_message,
        'success': success,
        'filter': f,
        'action': action,
        'layout': layout,
        'pk': pk,
        'percentages': percentages,
        'fios_doers': fios_doers
    }
    return render(request, r"schedulerfio/schedulerfio.html", context=context)


# авторизация пользователей
class LoginUser(LoginView):  # TODO перенести в service
    form_class = AuthenticationForm
    template_name = 'scheduler/login.html'

    def get_success_url(self):  # редирект после логина
        if 'admin' in self.request.user.username:
            return reverse_lazy('home')
        elif 'disp' in self.request.user.username:
            return reverse_lazy('scheduler')
        elif 'tehnolog' in self.request.user.username:
            return reverse_lazy('tehnolog')
        elif 'constructor' in self.request.user.username:
            return reverse_lazy('constructor')
        print(self.request.user.username)


def logout_user(request):  # разлогинивание пользователя  # TODO перенести в service
    logout(request)
    return redirect('login')


def show_workshop_scheme(request):  # TODO перенести в service
    """
    Загрузка планировки
    :param request:
    :return:
    """
    try:
        path_to_file = r"M:\Xranenie\ПТО\1 Екименко М.А\Планировка\Планирока участков(цех1, цех2, цех3)+РЦ+Расписание+Виды.xlsm"
        response = FileResponse(open(fr'{path_to_file}', 'rb'))
        response['X-Frame-Options'] = 'SAMEORIGIN'
        return response
    except FileNotFoundError as e:
        print(e)


@login_required(login_url="login")
def plan(request):
    """
    Планирование графика цеха и создание запросов на КД
    :param request:
    :return:
    """
    if str(request.user.username).strip()[:5] != "admin" and str(request.user.username[:4]).strip() != "disp":
        raise PermissionDenied
    group_id = TERMINAL_GROUP_ID  # тг группа

    # обновление процента готовности всех заказов
    # TODO модифицировать расчёт процента готовности всех заказов по взвешенной трудоёмкости
    #  сделать невозможным заполнять запрос с кириллицей
    get_all_done_rate()
    # график изделий
    workshop_schedule_fields = ('workshop', 'order', 'model_name', 'datetime_done', 'order_status', 'done_rate')
    workshop_schedule = (WorkshopSchedule.objects.values(*workshop_schedule_fields)
                         .exclude(datetime_done=None).exclude(order_status='не запланировано')
                         .order_by('datetime_done'))
    # фильтры в колонки графика
    f_w = get_filterset(data=request.GET, queryset=workshop_schedule, fields=workshop_schedule_fields, index=1)
    # фильтры в колонки заявок

    # форма запроса КД
    if request.method == 'POST':
        form_workshop_plan = SchedulerWorkshop(request.POST)
        if form_workshop_plan.is_valid():
            # заполнение графика цеха датой готовности и цехом
            try:
                # Планирование графика цеха
                # заполнение срока готовности, категория изделия, статус изделия, ФИО планировщика
                (WorkshopSchedule.objects.filter(model_order_query=form_workshop_plan.
                                                 cleaned_data['model_order_query'].model_order_query).update(
                    datetime_done=form_workshop_plan.cleaned_data['datetime_done'],
                    workshop=form_workshop_plan.cleaned_data['workshop'],
                    product_category=str(form_workshop_plan.cleaned_data['category']),
                    order_status='запланировано',
                    dispatcher_plan_ws_fio=f'{request.user.first_name} {request.user.last_name}'))
                # Заполнение данных СЗ, статус СЗ, ФИО планировщика, категория изделия,
                alert = 'Данные в график успешно занесены! '
                print('Данные в график успешно занесены!\n')
                # заполнение модели ShiftTask данными планирования цехов
                print(ShiftTask.objects.filter(model_order_query=form_workshop_plan.cleaned_data['model_order_query']))
                (ShiftTask.objects.filter(
                    model_order_query=form_workshop_plan.cleaned_data['model_order_query'].model_order_query)
                 .update(workshop=form_workshop_plan.cleaned_data['workshop'],
                         datetime_done=form_workshop_plan.cleaned_data['datetime_done'],
                         product_category=str(form_workshop_plan.cleaned_data['category']),
                         ))
                print('Данные сменного задания успешно занесены!')
                alert += 'Данные сменного задания успешно занесены.'
                context = {'form_workshop_plan': form_workshop_plan, 'alert': alert,
                           'workshop_schedule': workshop_schedule, 'filter_w': f_w, }
                # сообщение в группу
                success_group_message = (f"Заказ-модель: "
                                         f"{form_workshop_plan.cleaned_data['model_order_query'].model_order_query} "
                                         f"успешно запланирован на {form_workshop_plan.cleaned_data['datetime_done']}. "
                                         f"Запланировал: {request.user.first_name} {request.user.last_name}.")
                asyncio.run(terminal_message_to_id(to_id=group_id, text_message_to_id=success_group_message))
            except Exception as e:
                print(e, ' Ошибка запаси в базу SchedulerWorkshop')
                alert = f'Ошибка занесения данных.'
                context = {'form_workshop_plan': form_workshop_plan, 'workshop_schedule': workshop_schedule,
                           'alert': alert, 'filter_w': f_w, }
                return render(request, r"scheduler/scheduler.html", context=context)
        else:
            context = {'form_workshop_plan': form_workshop_plan, 'workshop_schedule': workshop_schedule,
                       'filter_w': f_w, }
    else:
        # чистые формы для первого запуска
        form_workshop_plan = SchedulerWorkshop()
        context = {'form_workshop_plan': form_workshop_plan, 'workshop_schedule': workshop_schedule,
                   'filter_w': f_w}
    return render(request, r"scheduler/plan.html", context=context)


@login_required(login_url="login")
def test_scheduler(request):
    """
        Планирование графика цеха и создание запросов на КД
        :param request:
        :return:
        """
    if str(request.user.username).strip()[:5] != "admin" and str(request.user.username[:4]).strip() != "disp":
        raise PermissionDenied
    group_id = TERMINAL_GROUP_ID  # тг группа

    # обновление процента готовности всех заказов
    # TODO модифицировать расчёт процента готовности всех заказов по взвешенной трудоёмкости
    #  сделать невозможным заполнять запрос с кириллицей
    # перечень запросов на КД
    td_queries_fields = ['model_order_query', 'query_prior', 'td_status', 'order_status', 'sz']  # поля таблицы
    td_queries = (WorkshopSchedule.objects.values(*td_queries_fields).exclude(order_status='завершено'))
    filter_fields = [
        'model_order_query',
        'query_prior',
        'td_status',
        'order_status',
    ]
    f_q = get_filterset(data=request.GET, queryset=td_queries, fields=filter_fields, index=2)
    sz_shift_tasks = ShiftTask.objects.values(
        "id", "workpiece", 'model_order_query'
    ).filter(
        st_status="не запланировано"
    )

    # форма запроса КД
    form_query_draw = QueryDraw()
    form_plan_bid = PlanBid()
    if request.method == 'POST':
        form_workshop_plan = SchedulerWorkshop(request.POST)
        if form_workshop_plan.is_valid():
            print(form_workshop_plan.cleaned_data)
            # заполнение графика цеха датой готовности и цехом
            try:
                # Планирование графика цеха
                # заполнение срока готовности, категория изделия, статус изделия, ФИО планировщика
                (WorkshopSchedule.objects.filter(model_order_query=form_workshop_plan.
                                                 cleaned_data['model_order_query'].model_order_query).update(
                    datetime_done=form_workshop_plan.cleaned_data['datetime_done'],
                    workshop=form_workshop_plan.cleaned_data['workshop'],
                    product_category=str(form_workshop_plan.cleaned_data['category']),
                    order_status='запланировано',
                    dispatcher_plan_ws_fio=f'{request.user.first_name} {request.user.last_name}'))
                # Заполнение данных СЗ, статус СЗ, ФИО планировщика, категория изделия,
                alert = 'Данные в график успешно занесены! '
                print('Данные в график успешно занесены!\n')
                # заполнение модели ShiftTask данными планирования цехов
                print(ShiftTask.objects.filter(model_order_query=form_workshop_plan.cleaned_data['model_order_query']))
                (ShiftTask.objects.filter(
                    model_order_query=form_workshop_plan.cleaned_data['model_order_query'].model_order_query)
                 .update(workshop=form_workshop_plan.cleaned_data['workshop'],
                         datetime_done=form_workshop_plan.cleaned_data['datetime_done'],
                         product_category=str(form_workshop_plan.cleaned_data['category']),
                         ))
                print('Данные сменного задания успешно занесены!')
                alert += 'Данные сменного задания успешно занесены.'
                context = {'form_workshop_plan': form_workshop_plan, 'td_queries': td_queries, 'alert': alert,
                           'form_query_draw': form_query_draw, 'filter_q': f_q}
                # сообщение в группу
                success_group_message = (f"Заказ-модель: "
                                         f"{form_workshop_plan.cleaned_data['model_order_query'].model_order_query} "
                                         f"успешно запланирован на {form_workshop_plan.cleaned_data['datetime_done']}. "
                                         f"Запланировал: {request.user.first_name} {request.user.last_name}.")
                asyncio.run(terminal_message_to_id(to_id=group_id, text_message_to_id=success_group_message))
            except Exception as e:
                print(e, ' Ошибка запаси в базу SchedulerWorkshop')
                alert = f'Ошибка занесения данных.'
                context = {'form_workshop_plan': form_workshop_plan, 'td_queries': td_queries,
                           'form_query_draw': form_query_draw, 'alert': alert, 'filter_q': f_q}
                return render(request, r"scheduler/scheduler.html", context=context)
        else:
            context = {'form_workshop_plan': form_workshop_plan, 'td_queries': td_queries,
                       'form_query_draw': form_query_draw, 'filter_q': f_q}
    else:
        # чистые формы для первого запуска
        form_workshop_plan = SchedulerWorkshop()
        context = {'form_workshop_plan': form_workshop_plan, 'td_queries': td_queries,
                   'form_query_draw': form_query_draw, 'filter_q': f_q, 'form_plan_bid': form_plan_bid,
                   'sz_st': sz_shift_tasks}
    return render(request, r"scheduler/test_scheduler.html", context=context)


def shift_tasks_reports(request, start: str = "", end: str = ""):  # TODO перенести в service
    """
    Загружает файл отчета по сменным заданиям
    :param request:
    :param start: с даты (дата распределения)
    :param end: по дату (дата распределения)
    :return: excel-файл
    """
    start, end = get_start_end_st_report(start, end)
    exel_file = create_shift_task_report(start, end)
    return FileResponse(open(exel_file, 'rb'))


def shift_tasks_auto_report():  # TODO перенести в service
    """
    Отправляет отчет по сменным заданиям на электронную почту и в папку O:/Расчет эффективности/Отчёты по СЗ
    """
    start = make_aware(datetime.datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0))
    end = make_aware(datetime.datetime.now())
    exel_file = create_shift_task_report(start, end)
    shutil.copy(exel_file, os.path.join(r"O:\Расчет эффективности\Отчёты по СЗ", os.path.basename(exel_file)))
    email = EmailMessage(
        f"Отчет {start.strftime('%d.%m.%Y')}-{end.strftime('%d.%m.%Y')}",
        f"Отчет {start.strftime('%d.%m.%Y')}-{end.strftime('%d.%m.%Y')}",
        "omzit-report@yandex.ru",
        [
            "alex4ekalovets@gmail.com",
            "pdo02@omzit.ru",
            "pdo06@omzit.ru",
            "pdo09@omzit.ru",
            "e.savchenko@omzit.ru",
            "PVB@omzit.ru",
            "m.ekimenko@omzit.ru"
        ],
        [],
    )
    email.attach_file(exel_file)
    email.send()


def create_shift_task_report(start, end):  # TODO перенести в service
    """
    Создает excel-файл отчета по сменным заданиям в папке xslx в корне проекта
    :param start: с даты (дата распределения)
    :param end: по дату (дата распределения)
    :return:
    """
    # Формируем имена столбцов для полного отчета из аттрибута модели verbose_name
    verbose_names = dict()
    for field in ShiftTask._meta.get_fields():
        if hasattr(field, "verbose_name"):
            verbose_names[field.name] = field.verbose_name
        else:
            verbose_names[field.name] = field.name

    queryset = ShiftTask.objects.exclude(
        fio_doer="не распределено"
    ).order_by("datetime_assign_wp")

    fields_1C_report = (
        "pk",  # №
        "model_name",  # модель
        "order",  # заказ
        "op_number",  # № Операции
        "op_name_full",  # Операция
        "fio_doer",  # Исполнители
        "decision_time",  # Дата готовности
        "st_status",  # статус СЗ
    )
    fields_disp_report = (
        "pk",  # №
        "model_name",  # модель
        "order",  # заказ
        "ws_number",  # РЦ
        "op_number",  # № Операции
        "op_name_full",  # Операция
        "fio_doer",  # Исполнители
        "datetime_assign_wp",  # Дата распределения
        "datetime_job_start",  # Дата начала
        "decision_time",  # Дата окончания
        "job_duration",  # Длительность работы
        "norm_tech",  # Технологическая норма
        "st_status",  # Статус СЗ
        "master_finish_wp",  # Мастер
        "otk_decision",  # Контролер
    )

    # Определяем путь к excel файлу шаблона
    exel_file_src = BASE_DIR / "ReportTemplate.xlsx"
    # Формируем название нового файла
    new_file_name = (f"{datetime.datetime.now().strftime('%Y.%m.%d')} report "
                     f"{start.strftime('%d.%m.%Y')}-{end.strftime('%d.%m.%Y')}.xlsx")
    # Создаем папку для хранения отчетов
    if not os.path.exists(BASE_DIR / "xlsx"):
        os.mkdir(BASE_DIR / "xlsx")
    # Формируем путь к новому файлу
    exel_file_dst = BASE_DIR / "xlsx" / new_file_name
    # Копируем шаблон в новый файл отчета
    shutil.copy(exel_file_src, exel_file_dst)

    # Формируем отчет
    ex_wb = openpyxl.load_workbook(exel_file_src, data_only=True)
    sheets_reports = {
        "Отчет для 1С": queryset.values(*fields_1C_report).filter(
            datetime_assign_wp__gte=start,
            datetime_assign_wp__lte=end
        ),
        "Отчет для диспетчера": queryset.values(*fields_disp_report).filter(
            datetime_assign_wp__gte=start,
            datetime_assign_wp__lte=end
        ),
        "Полный отчет": queryset.values(*verbose_names)
    }
    for sheet_name in sheets_reports:
        ex_sh = ex_wb[sheet_name]
        report = sheets_reports[sheet_name]
        if report:
            # Для полного отчета создаем шапку из verbose_name
            if sheet_name == "Полный отчет":
                for i, key in enumerate(report[0]):
                    ex_sh.cell(row=1, column=i + 1).value = verbose_names[key]
            # Заполняем строки данными
            for i, row in enumerate(report):
                for j, key in enumerate(row):
                    cell = ex_sh.cell(row=i + 2, column=j + 1)
                    try:
                        row[key] = make_naive(row[key]).strftime('%Y.%m.%d %H:%M:%S')
                    except Exception:
                        pass
                    cell.value = row[key]
            ex_wb.save(exel_file_dst)
    return exel_file_dst


def shift_tasks_report_view(request, start: str = "", end: str = ""):
    """
    Просмотр отчета по сменным заданиям за выбранный период
    :param request:
    :param start: с даты (Дата распределения)
    :param end: по дату (Дата распределения)
    """
    start, end = get_start_end_st_report(start, end)
    shift_task_fields = (
        'id', 'workshop', 'order', 'model_name', 'datetime_done', 'ws_number',
        'op_number', 'op_name_full', 'norm_tech', 'fio_doer', "datetime_assign_wp", 'st_status',
    )

    workplace_schedule = ShiftTask.objects.values(
        *shift_task_fields
    ).exclude(
        fio_doer="не распределено"
    ).filter(
        datetime_assign_wp__gte=start,
        datetime_assign_wp__lte=end
    ).order_by("datetime_assign_wp")

    f = get_filterset(data=request.GET, queryset=workplace_schedule, fields=shift_task_fields)
    context = {
        'workplace_schedule': workplace_schedule,
        'filter': f,
    }
    return render(request, fr"schedulerwp/view_report.html", context=context)


def get_start_end_st_report(start: str, end: str) -> Tuple:  # TODO перенести в service
    """
    Преобразует полученные от пользователя строки с датами или null в дату и время
    :param start: с даты (Дата распределения)
    :param end: по дату (Дата распределения)
    :return: дату начала, дату окончания формирования отчета
    """
    if start == "null":
        start = make_aware(datetime.datetime(year=1990, month=1, day=1, hour=0, minute=0, second=0, microsecond=0))
    else:
        start = make_aware(datetime.datetime.strptime(start, "%d.%m.%Y"))
    if end == "null":
        end = make_aware(datetime.datetime.now())
    else:
        end = make_aware(datetime.datetime.strptime(end, "%d.%m.%Y").replace(hour=23, minute=59, second=59))
    return start, end


def create_specification(request):
    """
    Рабочее место для создания заявки на детали по спецификации
    """
    alert = ""
    spec = dict()
    draw_form = CdwChoiceForm()  # Форма выбора файлов
    send_form = SendSZForm()  # Форма отправки служебной записки
    shared_folder = r"\\omzit\Shared\Temp\cdwr"

    ip = request.META.get('REMOTE_ADDR')

    # Проверяем, есть ли запущенный процесс загрузки чертежей для данного ip
    pid = SPEC_CREATION_PROCESS.get(ip)
    if pid and psutil.pid_exists(pid):
        alert = "Выполняется формирование спецификации..."
        context = {'spec': spec, "send_form": send_form, 'alert': alert, "draw_form": draw_form}
        return render(request, r"scheduler/specification.html", context=context)

    thr = SPEC_CREATION_THREAD.get(ip)
    if thr and thr.is_alive():
        alert = "Выполняется формирование спецификации..."
        context = {'spec': spec, "send_form": send_form, 'alert': alert, "draw_form": draw_form}
        return render(request, r"scheduler/specification.html", context=context)
    form_submit = request.POST.get("form", "")
    if request.method == "POST" and form_submit == "clear_form":  # очистка таблицы
        if os.path.exists(shared_folder):
            try:
                shutil.rmtree(shared_folder)
            except Exception as ex:
                print(f"Во время удаления папки {shared_folder} возникло исключение: {ex}")
    elif request.method == "POST" and form_submit == "td_kd_form":  # добавление файлов
        draw_form = CdwChoiceForm(request.POST, request.FILES)
        files = []
        if draw_form.is_valid():
            files = dict(request.FILES).get("cdw_files")

            # Создаем директорию cdwr в папке M:\Temp
            try:
                os.mkdir(shared_folder)
            except Exception as ex:
                print(fr"При попытке создания папки {shared_folder} вызвано исключение: {ex}")
            if os.path.exists(shared_folder):
                files_path = shared_folder
                # client = None
                client = connect_to_client(ip)
                if client:
                    scenario = "КОМПАС на клиенте"
                else:
                    scenario = "КОМПАС на сервере"
            else:  # Если нет доступа к M:\Temp, то создаем папку на сервере
                files_path = BASE_DIR / "cdw"
                if not os.path.exists(files_path):
                    os.mkdir(files_path)
                scenario = "КОМПАС на сервере"

            # Загружаем файлы, добавляем полные пути в список files_paths
            files_paths = []
            for file in files:
                filename = str(file)
                try:
                    handle_uploaded_file(f=file, filename=filename, path=files_path)
                    files_paths.append(os.path.join(files_path, filename))
                except Exception as ex:
                    print(f"При загрузке файлов в {files_path} возникло исключение: {ex}")
                    files_paths = []

            print(scenario)
            # Обработка сценария, когда КОМПАС установлен на сервере
            if scenario == "КОМПАС на сервере":
                pid = get_specifications_server(files_paths)
                SPEC_CREATION_PROCESS[ip] = pid
            elif scenario == "КОМПАС на клиенте":
                thread = threading.Thread(target=get_specifications_ssh, args=(client, files_paths))
                SPEC_CREATION_THREAD[ip] = thread
                thread.start()

            if files_paths:
                alert = "Выполняется формирование спецификации..."
            else:
                alert = "Ошибка формирования спецификации!"
        context = {'spec': spec, 'draw_form': draw_form, 'alert': alert, "files": files, "send_form": send_form}
        return render(request, r"scheduler/specification.html", context=context)
    else:  # метод GET
        # Загружаем файл спецификации из specification.json
        try:
            if os.path.exists(shared_folder):
                files_path = shared_folder
            else:
                files_path = BASE_DIR / "cdw"
            json_path = os.path.join(files_path, "specification.json")
            with open(json_path, 'r') as json_file:
                spec = json.load(json_file)
        except Exception:
            print("Ошибка получения файла спецификации")
        rows = []  # Все строки по всем чертежам
        names = []  # Наименования для заполнения select
        draw_names = dict()  # Соответствие наименований чертежам
        draws = set()  # Чертежи для заполнения select
        if spec:
            spec.pop("columns")
            for key, value in spec.items():
                draw_names[key] = []
                for row in value:
                    draw_names[key].append(row["Наименование"])
                    row["Чертеж"] = key
                    draws.add(key)
                    names.append(row["Наименование"])
                rows.extend(value)
    context = {'draw_names': draw_names, 'rows': rows, "names": names, "draws": draws,
               "send_form": send_form, "draw_form": draw_form, 'alert': alert}
    return render(request, r"scheduler/specification.html", context=context)


def create_shift_tasks_from_spec(request):
    """
    Создает сменные задания из спецификации
    """
    pdf_sz_filename = BASE_DIR / "example.pdf"
    if request.method == "POST":

        # Удаляем файлы спецификаций
        json_path_1 = BASE_DIR / "specification.json"
        json_path_2 = r"\\omzit\Shared\Temp\cdwr\specification.json"
        if os.path.exists(json_path_1):
            os.remove(json_path_1)
        elif os.path.exists(json_path_2):
            os.remove(json_path_2)

        json_data = request.body
        data = json.loads(json_data)

        if data["products"] and all(value for value in data["sz"].values()):
            data["sz"]["author"] = request.user.first_name
            # Создание служебной записки в pdf
            create_pdf_sz(data=data, filename=pdf_sz_filename)

            model_name = str(int(datetime.datetime.now().timestamp()))
            order = data["sz"]["sz_number"]
            model_order_query = f"{order}_{model_name}"
            WorkshopSchedule.objects.create(
                model_name=model_name,
                order=order,
                model_order_query=model_order_query,
                sz=data["sz"],
                td_status="завершено",
            )
            shift_tasks = []
            for product in data["products"]:
                product['count'] = int(product['count']) if product['count'].isdigit() else 0
                shift_tasks.append(ShiftTask(
                    model_name=model_name,
                    order=order,
                    model_order_query=model_order_query,
                    workpiece=product,
                ))
            ShiftTask.objects.bulk_create(shift_tasks)
            return FileResponse(open(pdf_sz_filename, 'rb'))
        else:
            return JsonResponse({"STATUS": "No data"})


def confirm_sz_planning(request):
    alert = ""
    if request.method == "POST":
        json_data = request.body
        data = json.loads(json_data)
        if data.values():
            model_order_query = f"{data['newOrder']}_{data['newModel']}"
            ws = WorkshopSchedule.objects.filter(model_order_query=data["orderModel"])
            shift_tasks = ShiftTask.objects.filter(
                model_order_query=data["orderModel"],
                id__in=list(map(int, data["st"].keys()))
            )
            try:
                ws.update(
                    datetime_done=make_aware(datetime.datetime.strptime(data['dateDone'], "%d.%m.%Y")),
                    workshop=data['workshop'],
                    product_category=data["category"],
                    order_status='запланировано',
                    dispatcher_plan_ws_fio=f'{request.user.first_name} {request.user.last_name}',
                    model_order_query=model_order_query,
                    model_name=data['newModel'],
                    order=data['newOrder'],
                )
            except IntegrityError as ex:
                alert = f"Ошибка! Заказ-модель {model_order_query} уже существует!"
                print(f"При обновлении записей {ws} возникло исключение: {ex}")
                return JsonResponse({"STATUS": alert})

            for st in shift_tasks:
                attrs = {
                    "datetime_done": make_aware(datetime.datetime.strptime(data['dateDone'], "%d.%m.%Y")),
                    "workshop": data['workshop'],
                    "product_category": data["category"],
                    "order_status": 'запланировано',
                    "model_order_query": model_order_query,
                    "model_name": data['newModel'],
                    "order": data['newOrder'],
                }
                if data['st'][str(st.pk)] == "Плазма":
                    attrs.update({
                        'ws_number': "",
                        'ws_name': "Плазма",
                        "st_status": "раскладка"
                    })
                else:
                    attrs.update({
                        'ws_number': data['st'][str(st.pk)],
                        "st_status": "запланировано"
                    })

                for attr, value in attrs.items():
                    setattr(st, attr, value)
                st.save()
            return JsonResponse({"STATUS": "OK"})
        else:
            return JsonResponse({"STATUS": alert})

# @login_required(login_url="login")
# def report(request, workshop):
#     """
#     # TODO ФУНКЦИОНАЛ ОТЧЁТОВ законсервировано до момента когда кому-то понадобится
#     Отображение формы для заполнения ежедневного отчёта. Построение и сохранение графика отчётов по цехам.
#     :param workshop: Номер цеха
#     :param request:
#     :return:
#     """
#     if str(request.user.username).strip()[:5] != "admin" and str(request.user.username[:4]).strip() != "disp":
#         raise PermissionDenied
#     # days_report_create()  # запускается по расписанию
#     today = datetime.datetime.now()  # сегодня
#     # today = datetime.datetime(year=2023, month=11, day=25)  # тестовый сегодня - произвольная дата
#     # получение даты из секретного поля через GET # TODO закрыть уязвимость
#     if request.GET.get('a', None):
#         print(f"{request.GET.get('a', None)=}")
#         today = datetime.datetime.strptime(request.GET.get('a'), "%Y-%m-%d")
#     # данные дат
#     yesterday = today - datetime.timedelta(days=1)  # вчера
#     start_date = yesterday.replace(day=1)  # первый день текущего месяца
#     _, last_day = calendar.monthrange(yesterday.year, yesterday.month)  # последний день текущего месяца
#     end_date = datetime.datetime(year=yesterday.year, month=yesterday.month, day=last_day)  # последняя дата месяца
#     month_plan = MonthPlans.objects.get(workshop=workshop, month_plan=start_date)  # план цеха на месяц
#     # отображение интервалов на для графика производства
#     report_days_for_plot = (DailyReport.objects.filter(workshop=workshop, calendar_day__lte=yesterday,
#                                                        calendar_day__gte=start_date)
#                             .order_by('calendar_day'))
#     # отображение интервалов на для окна планирования
#     report_days_for_plans = (DailyReport.objects.filter(workshop=workshop,
#                                                         calendar_day__gte=start_date).order_by('calendar_day'))
#     # отображение интервалов на главной странице
#     report_days = (DailyReport.objects.filter(calendar_day__gte=today - datetime.timedelta(days=3),
#                                               calendar_day__lte=today + datetime.timedelta(days=3),
#                                               workshop=workshop)
#                    .order_by('calendar_day'))
#
#     if request.method == 'POST':
#         report_form = DailyReportForm(request.POST)
#         hidden_field = PlanResortHiddenForm()
#         if report_form.is_valid():
#             # данные за вчера
#             yesterday_data = DailyReport.objects.select_related('month_plan_data').get(calendar_day=yesterday,
#                                                                                        workshop=workshop)
#             print(f'{report_form.cleaned_data=}')
#             try:
#                 if yesterday.month == (yesterday - datetime.timedelta(days=1)).month:
#                     previous_day_data = DailyReport.objects.get(
#                         calendar_day=(yesterday - datetime.timedelta(days=1)).date(), workshop=workshop)
#                     fact_sum = previous_day_data.fact_sum + report_form.cleaned_data['day_fact']
#                     plan_sum = previous_day_data.plan_sum + yesterday_data.day_plan
#                 else:
#                     fact_sum = report_form.cleaned_data['day_fact']
#                     plan_sum = yesterday_data.day_plan
#             except Exception as e:
#                 print("Ошибка получения previous_day_data: ", e)
#                 fact_sum = report_form.cleaned_data['day_fact']
#                 plan_sum = yesterday_data.day_plan
#             # расчёт плановых показателей
#             day_plan_rate = 100 * report_form.cleaned_data['day_fact'] / yesterday_data.day_plan
#             # plan_done_rate = 100 * fact_sum / yesterday_data.plan_sum
#             plan_done_rate = 100 * yesterday_data.plan_sum / yesterday_data.month_plan_data.month_plan_amount
#             fact_done_rate = 100 * fact_sum / yesterday_data.month_plan_data.month_plan_amount
#             plan_loos_rate = fact_done_rate - plan_done_rate
#             # чтение json
#             json_file_path = r'M:\Xranenie\ПТО\1 Екименко М.А'
#             json_filename = os.path.join(json_file_path, "all_indicators.json")
#             with open(json_filename, 'r') as json_file:
#                 line = json_file.readline()
#                 while line:
#                     if str(yesterday.date()) in line:
#                         json_raw_file_content = json.loads(line)
#                         json_file_content = json_raw_file_content[str(yesterday.date())]
#                         break
#                     else:
#                         json_file_content = dict()
#                     line = json_file.readline()
#             # получение данных ОТК, ОТПБ и дуги
#             day_fails = json_file_content.get(f'c{workshop}_fails_day', [0])[0]
#             day_save_violations = json_file_content.get(f'c{workshop}_un_save_day', [0])[0]
#             # значение сумм нарушений по умолчанию
#             day_fails_sum = DailyReport.objects.filter(workshop=workshop,
#                                                        calendar_day__gte=start_date,
#                                                        calendar_day__lte=end_date).aggregate(
#                 res_sum=Sum('day_fails'))['res_sum']
#             day_violations_sum = DailyReport.objects.filter(workshop=workshop, calendar_day__gte=start_date,
#                                                             calendar_day__lte=end_date).aggregate(
#                 res_sum=Sum('day_save_violations'))['res_sum']
#             day_fails_str = json_file_content.get(f'c{workshop}_fails_result', f'{0}/{day_fails_sum}')
#             day_save_violations_str = json_file_content.get(f'c{workshop}_un_save_result', f'{0}/{day_violations_sum}')
#             # получение данных дуги
#             arc_str = json_file_content.get(f'ceh{workshop}_arc', '0')
#             # обновление БД
#             record_object = DailyReport.objects.filter(calendar_day=yesterday.date(), workshop=workshop)
#             record_object.update(fact_sum=fact_sum, plan_sum=plan_sum,
#                                  day_fact=report_form.cleaned_data['day_fact'],
#                                  day_plan_rate=day_plan_rate, plan_done_rate=plan_done_rate,
#                                  fact_done_rate=fact_done_rate, plan_loos_rate=plan_loos_rate,
#                                  # aver_fact=aver_fact['average'],
#                                  day_fails=day_fails,
#                                  day_save_violations=day_save_violations,
#                                  personal_total=report_form.cleaned_data['personal_total'],
#                                  personal_shift=report_form.cleaned_data['personal_shift'],
#                                  personal_total_welders=report_form.cleaned_data['personal_total_welders'],
#                                  personal_shift_welders=report_form.cleaned_data['personal_shift_welders'],
#                                  personal_night_welders=report_form.cleaned_data['personal_night_welders'],
#                                  personal_total_locksmiths=report_form.cleaned_data['personal_total_locksmiths'],
#                                  personal_shift_locksmiths=report_form.cleaned_data['personal_shift_locksmiths'],
#                                  personal_night_locksmiths=report_form.cleaned_data['personal_night_locksmiths'],
#                                  personal_total_painters=report_form.cleaned_data['personal_total_painters'],
#                                  personal_shift_painters=report_form.cleaned_data['personal_shift_painters'],
#                                  personal_night_painters=report_form.cleaned_data['personal_night_painters'],
#                                  personal_total_turners=report_form.cleaned_data['personal_total_turners'],
#                                  personal_shift_turners=report_form.cleaned_data['personal_shift_turners'],
#                                  personal_night_turners=report_form.cleaned_data['personal_night_turners'],
#                                  )
#             # отображение интервалов на для графика производства
#             report_days_for_plot = (DailyReport.objects.filter(workshop=workshop, calendar_day__lte=yesterday,
#                                                                calendar_day__gte=start_date)
#                                     .order_by('calendar_day'))
#             # отображение интервалов на для окна планирования
#             report_days_for_plans = (DailyReport.objects.filter(workshop=workshop,
#                                                                 calendar_day__gte=start_date).order_by('calendar_day'))
#             # отображение интервалов на главной странице
#             report_days = (DailyReport.objects.filter(calendar_day__gte=today - datetime.timedelta(days=3),
#                                                       calendar_day__lte=today + datetime.timedelta(days=3),
#                                                       workshop=workshop)
#                            .order_by('calendar_day'))
#             # обновление словаря для импорта в json для передачи данных монитору
#             # json_raw_file_content[str(yesterday.date())].update(
#             #     [(f'c{workshop}_day_plan_rate', float(day_plan_rate)),
#             #      (f'c{workshop}_plan_done_rate', float(plan_done_rate)),
#             #      (f'c{workshop}_plan_loos_rate', float(plan_loos_rate))])
#             # обновление json для
#             # with open(json_filename, 'w') as json_file:
#             #     json.dump(json_raw_file_content, json_file)
#             # формирование отчёта
#             aver_fact = report_days_for_plot.aggregate(average=Avg('day_plan_rate'))  # средний факт
#             record_object.update(aver_fact=aver_fact['average'])  # обновление средней
#             plot_lists = report_days_for_plot.values('calendar_day', 'day_plan_rate')
#             days_list = []
#             fact_list = []
#             for plot_list in plot_lists:
#                 days_list.append(plot_list['calendar_day'].strftime('%d.%m'))
#                 fact_list.append(float(plot_list['day_plan_rate']))
#             # объединение отчётов
#             report_merger(month=yesterday.month, merge_dir=os.path.join(BASE_DIR, r'scheduler\static\scheduler\pdf'))
#
#             # формирование графика
#             make_workshop_plan_plot(workshop=workshop, days_list=days_list, fact_list=fact_list,
#                                     aver_fact=aver_fact['average'],
#                                     start_day=start_date, end_day=end_date, yesterday=yesterday)
#             # формирование результирующего отчёта pdf
#             filename = f'plan{workshop}-{yesterday.month}.pdf'
#             image_path = rf'scheduler\static\scheduler\jpg\plan{workshop}-{yesterday.month}.jpg'
#             output_dir = rf'scheduler\static\scheduler\pdf'
#             personal_data = DailyReport.objects.get(workshop=workshop, calendar_day=yesterday)
#             header_text = f"Данные для цеха {workshop} на {yesterday.strftime('%d.%m.%y')}"
#             table_data = (('Персонал', 'Cварщики', 'Слесаря', 'Маляры', 'Токари', 'Средняя дуга', 'Случаев брака',
#                            'Нарушений ОТ'),
#                           (f'{personal_data.personal_total} | {personal_data.personal_shift}',
#                            f'{personal_data.personal_total_welders} | {personal_data.personal_shift_welders} | '
#                            f'{personal_data.personal_night_welders}',
#                            f'{personal_data.personal_total_locksmiths} | {personal_data.personal_shift_locksmiths} | '
#                            f'{personal_data.personal_night_locksmiths}',
#                            f'{personal_data.personal_total_painters} | {personal_data.personal_shift_painters} | '
#                            f'{personal_data.personal_night_painters}',
#                            f'{personal_data.personal_total_turners} | {personal_data.personal_shift_turners} | '
#                            f'{personal_data.personal_night_turners}',
#                            arc_str, day_fails_str, day_save_violations_str))
#             footer_text = ["Обозначение для ячеек персонала: 'ВСЕГО | ВЫХОД | ВЫХОД НОЧЬ",
#                            "Обозначение для средней дуги: 'Дуга, ч / количество аппаратов",
#                            "Обозначение для брака и ОТ, случаев на дату / всего с начала месяца"]
#             create_pdf_report(filename=filename, image_path=image_path, output_dir=output_dir, table_data=table_data,
#                               header_text=header_text, footer_text=footer_text)
#             context = {'report_form': report_form, 'report_days': report_days, 'yesterday': yesterday.date(),
#                        'workshop': workshop, 'report_days_for_plot': report_days_for_plot, 'month_plan': month_plan,
#                        'report_days_for_plans': report_days_for_plans, 'hidden_field': hidden_field}
#             # print(report_form.cleaned_data)
#             return render(request, r"scheduler/report.html", context=context)
#         else:
#             print('not_valid_form')
#     else:
#         report_form = DailyReportForm()
#         hidden_field = PlanResortHiddenForm()
#     context = {'report_form': report_form, 'report_days': report_days, 'yesterday': yesterday.date(),
#                'workshop': workshop, 'report_days_for_plot': report_days_for_plot, 'month_plan': month_plan,
#                'report_days_for_plans': report_days_for_plans, 'hidden_field': hidden_field}
#     return render(request, r"scheduler/report.html", context=context)
#
#
# def plan_resort(request):
#     """
#     Обработка заполнения плана на каждый день по месяцам
#     :param request:
#     :return:
#     """
#     if request.method == 'POST':
#         hidden_field = PlanResortHiddenForm(request.POST)
#         if hidden_field.is_valid():
#             # получение данных из скрытого поля
#             raw_data = OrderedDict(json.loads(hidden_field.cleaned_data['day_plan_sum']))  # словарь результатов
#             workshop = raw_data.pop('workshop')
#             new_plan_sum = 0  # новая сумма плана
#             # обновление БД
#             records = []
#             first_day = list(raw_data.keys())[0]  # первый день месяца
#             print(first_day)
#             for _date, _plan in raw_data.items():
#                 formatted_date = datetime.datetime.strptime(_date, "%d.%m.%Y").date()
#                 formatted_first_day = datetime.datetime.strptime(first_day, "%d.%m.%Y").date()
#                 record = DailyReport.objects.get(calendar_day=formatted_date, workshop=workshop)
#                 new_plan_sum += _plan
#                 record.day_plan = _plan
#                 records.append(record)
#             # обновление ежедневного плана
#             DailyReport.objects.bulk_update(records, ['day_plan'])
#             # обновление полного плана на месяц
#             MonthPlans.objects.filter(workshop=workshop,
#                                       month_plan=formatted_first_day).update(month_plan_amount=round(new_plan_sum, 0))
#         else:
#             print('not valid form')
#
#     return redirect('report', workshop=workshop)  # обновление страницы при успехе
