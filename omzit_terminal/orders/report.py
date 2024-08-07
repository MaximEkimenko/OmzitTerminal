import pandas as pd
from pathlib import Path
from datetime import datetime


from django.utils.timezone import make_naive
from openpyxl import Workbook

from controller.utils.utils import get_model_verbose_names
from m_logger_settings import logger
from omzit_terminal.settings import BASE_DIR
from orders.models import Orders


from orders.utils.utils import orders_to_dict, get_order_verbose_names, ORDER_REPORT_COLUMNS


def create_order_report():
    wb = Workbook()
    ws = wb.active
    ws.title = "Заявки на ремонт"

    qs = Orders.fresh_orders()
    qs = orders_to_dict(qs, ORDER_REPORT_COLUMNS)
    verbose_header = get_model_verbose_names(Orders)
    verbose_header.update({"id": "№"})

    for i, column in enumerate(ORDER_REPORT_COLUMNS):
        ws.cell(row=1, column=i + 1).value = verbose_header[column]

    for i, row in enumerate(qs):
        for j, key in enumerate(row):
            cell = ws.cell(row=i + 2, column=j + 1)
            try:
                row[key] = make_naive(row[key])
            except Exception as e:
                # logger.exception(e)
                pass
            try:
                cell.value = row[key]
            except ValueError as e:
                logger.error(f"Ошибка при конвертации значения {row[key]} по ключу {key}")
                logger.exception(e)
                cell.value = "error"

    x = datetime.now()
    filename = f"Заявки на ремонт {x.day:02d}_{x.month:02d}_{x.year:04d} {x.hour:02d}_{x.minute:02d}_{x.second:02d}.xlsx"
    exel_file_dst = Path(BASE_DIR).joinpath("xlsx")
    abspath = exel_file_dst.joinpath(filename)
    wb.save(abspath)
    return abspath
