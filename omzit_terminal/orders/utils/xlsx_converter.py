from openpyxl import load_workbook, Workbook
from openpyxl.cell import Cell
from openpyxl.utils import rows_from_range
from pathlib import Path



def get_cell_border(cell: Cell) -> str:
    """
    Проходит по всем сторонам ячейки, и смотрит, есть ли видимые границы. Если есть,
    то формирует строку стиля с указанием имебщихся границ
    @param cell: Ячейка, взятая из экселя.
    @return Возвращает строку с указанием стилей для четырех сторон ячейки, рисовать на них границы или нет.
    """
    border_style = ""
    for direction in ["right", "left", "top", "bottom"]:
        b_s = getattr(cell.border, direction)
        # гранича есть и ее стиль отличается о дефолтного (то есть граница установлена пользователем)
        if b_s and b_s.style is not None:
            # мне нужно чтобы границы ячеек были как в остальных таблицах на сайте,
            # поэтому цвет применяю явно
            border_style += f"border-{direction}: solid 1px orange;"
        else:
            border_style += f"border-{direction}: none; "
    return border_style


def get_css_class_dict(tags: str, classes: dict) -> dict[str, str]:
    """
    Создает словарь, в котором тегу сопоставляется имя класса, полученное из внешнего словаря.
    Если во внешнем словаре искомого тега нет, то класс такому тегу не присавивается.
    @param tags: Список тегов, которые используются при формирвании html-таблицы.
    @param classes: Внешний словарь, где названиям тегов уже сопоставлены классы.
    Но некоторые теги в нем могут отсутствовать, это не приведет к ошибке.
    @return: Возвращает словарь, где каждому тегу сопоставлен фрагмент описания тега, где определяется класс.
    """
    class_dict = {}
    for tag in tags:
        c = classes.get(tag)
        class_dict[tag] = f' class="{c}"' if c else ""
    # мне нужно чтобы по умолчанию у таблицы был класс table
    if not class_dict["table"]:
        class_dict["table"] = ' class="table"'
    return class_dict


def tag_factory(css_class: dict | None) -> dict:
    """
    Создает словарь, где имени тега сопоставлен шаблон тега с определенным классом или стилем.
    В этот шаблон нужно только вставить содержимое. Например, в шаблон ряда можно вставить содержимое ячеек.
    @param css_class: Внешний словарь, где названиям тегов сопоставлены пользовательские классы.
    Отсутствовать. Тогда все ячейки будут формироваться на основе стилей, а не классов.
    @return: Словарь: тег - шаблона тега с определенным классом.
    """
    if css_class is None:
        css_class = {}
    tag_dict = {}
    tags = ["table", "thead", "tbody", "tr", "th", "td"]
    class_dict = get_css_class_dict(tags, css_class)
    for tag in tags:
        s_start = f'<{tag}{class_dict[tag]}>'
        s_end = f'</{tag}>'
        s = s_start + "{}" + s_end
        tag_dict[tag] = s
    return tag_dict


def xlsx_to_html(wb: Workbook,
                 worksheets: list[str] = None,
                 classes: dict[str, str] = None,
                 header: bool = False
                 ) -> dict[str, str]:
    """
    Конвертирует страницы эксель-файла в html-страницы.
    @param wb: Экселевский документ, из которого происходит экспорт
    @param worksheets: Список листов документа, которые нужно сконвертировать в HTML.
    Если он не указан, то конвертируются все листы документа.
    @param classes: Словарь, содержащий классы, для применения к html-тегам. Если отсутствует, то каждая ячейка
    форматируется индивидуально на основе свойств ячейки в экселе.
    @param header: Если True, то первая строка превращается в заголовок таблицы. Имеет смысл только при передаче
    словаря css-классов. У заголовка просто будет другой стиль. Больше ничего не изменится.
    @return: Словарь, в качестве ключа название листа документа, в качестве значения: сконвертированный в HTML лист.
    """

    use_classes = True if classes else False

    def make_cell(cell: Cell, tag: str, use_clsasses: bool) -> str:
        """
        Превращает ячейку эксель-документа в ячейку html-таблицы.
        @param cell: Ячейка из экселя.
        @param tag: Тэг td или th, в зависимости от того, в заголовке ячейка или в теле таблицы.
        @param use_clsasses: применять к тэгам классы стилей или форматировать каждую ячейку индивидуально
        с помошью параметра style.
        @return: Возвращает html-тег ячейки таблицы с определенным классом или индивидуальным стилем,
        скопированным с эксель ячейки. Главнаое свойство, которе копироуется из стиля - это границы ячейки.
        Индивидуальное применение границ позволяет рисовать красивые таблицы: именно так, как сони сделаны в экселе.
        """
        value = cell.value if cell.value is not None else ""
        cell_content = None

        if cell.fill.patternType is not None:
            # bgcolor = f"background-color: #{cell.fill.fgColor.rgb[2:]};"
            # решил заменить конкретным цветом под свою задачу
            bgcolor = f"background-color: rgba(100, 100, 100, 0.4);"
        else:
            bgcolor = f"background-color: none;"

        border_style = get_cell_border(cell)
        rowspan_attr = ""
        colspan_attr = ""
        sz = f"font-size: {int(cell.font.sz * 1.6)}px;" if cell.font.sz else ""
        bold = "font-weight: bold;" if cell.font.b else ""
        italic = "font-style: italic;" if cell.font.i else ""
        # если ячейка слитая, то ей надо указать, на сколько строк и стрлбщов она растягивается
        if cell in first_merged_cells:
            merge_cols = first_merged_cells[cell][0]
            merge_rows = first_merged_cells[cell][1]
            rowspan_attr = f' rowspan="{merge_rows}"'
            colspan_attr = f' colspan="{merge_cols}"'
        # игнорируем слитые ячейки (кроме первых), потому что выше все нужные уже растянули до правильного размера
        if cell not in excluded_cells:
            if use_clsasses:
                cell_content = tag_dict[tag].format(value)
            else:
                cell_content = (f'<{tag} {rowspan_attr}{colspan_attr} '
                                f'style="{sz} {bold} {italic} {bgcolor} {border_style}"'
                                f'>{value}</{tag}>')
        return cell_content

    if not worksheets:
        worksheets = wb.sheetnames
    converted_worksheets = {}
    for sheet_name in worksheets:
        ws = wb[sheet_name]
        # Ячейки, исключенные из построения таблицы. Это слитые ячейки, не несущие информации.
        # За исключением первой ячейки в диапазоне
        excluded_cells = set()
        for i in ws.merged_cells.ranges:
            for j in rows_from_range(i.coord):  # по рядам диапазона
                for mc in j:  # по ячейкам в ряду диапазона
                    excluded_cells.add(ws[mc])

        # словарь содержащий размеры объединенных ячеек
        first_merged_cells = {}
        for i in ws.merged_cells.ranges:
            ml = list(ws[i.coord])
            first_cell = ml[0][0]
            cols = len(ml[0])
            rows = len(ml)
            first_merged_cells[first_cell] = (cols, rows)

        # удаляем первые ячейки слитых диапазонов из множества исключенных ячеек
        for i in first_merged_cells:
            excluded_cells.remove(i)

        # построение HTML-документа
        tag_dict = tag_factory(classes)
        current_row = 1
        html_rows = []
        thdeader_text = ""
        # обрабатываем заголовок, если нужно
        if header:
            for row in ws.iter_rows(min_row=current_row, max_row=current_row):
                current_row += 1
                cells = []
                for cell in row:
                    cell_content = make_cell(cell, "th", use_classes)
                    if cell_content is not None:
                        cells.append(cell_content)
            html_row = tag_dict['tr'].format("".join(cells))
            thdeader_text = tag_dict['thead'].format(html_row) + "\n"

        for row_num, row in enumerate(ws.iter_rows(min_row=current_row)):
            cells = []
            for cell in row:
                cell_content = make_cell(cell, "td", use_classes)
                if cell_content is not None:
                    cells.append(cell_content)
            html_row = tag_dict['tr'].format("".join(cells))
            html_rows.append(html_row)
        tbody_text = tag_dict['tbody'].format("\n".join(html_rows))
        table_text = tag_dict['table'].format((thdeader_text + tbody_text))
        converted_worksheets[sheet_name] = table_text
    return converted_worksheets


def html_save(directory, converted_sheets: dict, style_file=None):
    """
    Сохраняет сконвертированные листы в файлы, предварительно обернув их в основные html-теги.
    Сохраняемые имеют имена листов обработанного эксель-документа.
    @param directory: Каталог, куда выгружать сконвертированные файлы.
    @param converted_sheets: Словарь, содержащий сконвертированные листы.
    @param style_file: Путь к файлу стилей. Если указан, то он добавляется в ссылку на стили.
    @return: Ничего не возвращает.
    """
    html_template = """
        <!DOCTYPE html>
        <html>
        <head>
            {0}
            <style>
            </style>
        </head>
        <body>
          {1}
        </body>
        </html>
        """
    style = ""
    if style_file:
        style = f'<link rel="stylesheet" type="text/css" href="{css_file}">'

    for sheet in converted_sheets:
        # сохранение в файл
        html_table = html_template.format(style, converted_sheets[sheet])
        output_file = Path.joinpath(directory, f"{sheet}.html")
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(html_table)


def prepare_output_directory(source_directory: str, output_dirmane: str) -> Path:
    """
    Создает и возвращает директорию для выгрузки созданных HTML-файлов.
    Если в директории уже есть файлы, они удаляются.
    @param source_directory: Путь к рабочей директории.
    @param output_dirmane: Имя поддиректории для выгрузки результатов.
    @return: Возвращает путь к созданной поддиректории для выгрузки результатов.
    """
    source_directory = Path(source_directory)
    output_dir: Path = source_directory.joinpath(output_dirmane)
    if not output_dir.exists():
        output_dir.mkdir()
    else:
        for i in output_dir.glob("*"):
            i.unlink()
    return output_dir


def get_workbook(xlsx_file: str) -> Workbook:
    """
    Принимает путь к файлу и возвращает открытый экселевский документ
    @param xlsx_file:  путь к файлу
    @return: экселевский документ (Workbook)
    """
    xlsx_file = Path(xlsx_file)
    wb = load_workbook(xlsx_file)
    return wb


# словарь css-классов, где каждому тегу сопоставлен свой класс
test_classes = {"table": "table_css",
                "thead": "thead_css",
                "tbody": "tbody_css",
                "tr": "tr_css",
                "th": "th_css",
                "td": "td_css",
                }

if __name__ == '__main__':
    xlsx_file_tst = r'D:\python_temp\convert_from_excel\пример для конвертации.xlsx'
    css_file = r'..\scheduler.css'

    output_direcory = prepare_output_directory(Path(xlsx_file_tst).parent, "результат")
    workbook = get_workbook(xlsx_file_tst)
    # каждая ячейка форматируется индивидуально
    converted_sheets = xlsx_to_html(workbook, header=True)
    # ко всем тегам применяются стили
    # converted_sheets = xlsx_to_html(workbook, header=True, classes=test_classes)
    # ссылка на файл со стилями привязывается отдельно при формировании файла, со словарем css-классов она не связана
    html_save(output_direcory, converted_sheets, css_file)
