import copy
from typing import Dict, List

import openpyxl
import re

import pandas as pd

from scheduler.models import ShiftTask, WorkshopSchedule
from tehnolog.services.pandas_utils import df_handler


def get_excel_data_pandas(data: Dict, exel_file: str, excel_list: str):
    operations = df_handler(pd.read_excel(exel_file, excel_list))
    data_list = []
    for index, row in operations.iterrows():
        data_list.append(
            {
                'model_name': data['model_name'],
                'order': data['order'],
                'model_order_query': data['model_order_query'],
                'excel_list_name': f'{excel_list}-{index}',  # инкремент для гарантированно уникальной записи
                'op_number': row['п/п'],
                'op_name': row['Наименование работ'],
                'ws_name': row['Рабочий центр'],
                'op_name_full': row['Наименование работ'] + '-' + row['Рабочий центр'],
                'ws_number': str(row['№ рабочего центра']),
                'norm_tech': row['Загрузка оборудования на 1 котел, часов'],
                'doers_tech': row['Численность, чел.'],
                'norm_calc': row['Загрузка оборудования на 1 котел, часов'],
                'draw_filename': row['Чертеж'],

            }
        )
    return data_list


def tech_data_get(model_order_query: str, exel_file: str, excel_list: str):
    """
    Функция получает технологические данные из файла exel_file по списку моделей изделия model_list,
    который равен списку имен листов файла исключая имена листов в списке exclusion_list и сохраняет в модели
    django ProductModel, TechData
    :param model_order_query: заказ-модель
    :param exel_file: имя файла excel
    :param excel_list: модель изделия для получения данных - лист книги excel
    :return: None
    """
    # модель, заказ
    order, model_name = model_order_query.split('_')
    # общие данные для всех записей shift_task (СЗ)
    common_data = {
        'model_name': model_name,
        'order': order,
        'model_order_query': model_order_query
    }
    # формирование данных для создания записей shift_task (СЗ)
    # data_list = get_excel_data(common_data, exel_file, excel_list)
    data_list = get_excel_data_pandas(common_data, exel_file, excel_list)

    def create_all_shift_tasks():  # TODO при рефакторинге перенести в services
        """
        Метод создает сменные задания
        :return:
        """
        ws = WorkshopSchedule.objects.get(model_order_query=model_order_query)
        add_data = {
            'datetime_done': ws.datetime_done,
            'workshop': ws.workshop,
            'product_category': ws.product_category
        }
        tasks = []
        for data in data_list:
            data.update(add_data)
            tasks.append(ShiftTask(**data))
        ShiftTask.objects.bulk_create(tasks)

    # заполнение shift_task
    shift_tasks = ShiftTask.objects.filter(model_order_query=model_order_query)  # существующие СЗ
    is_uploaded = True
    # если ранее не загружалось
    if not shift_tasks.exists():
        create_all_shift_tasks()
    # если статус запланировано
    elif all(st == 'не запланировано' for st in shift_tasks.values_list('st_status', flat=True)):
        shift_tasks.delete()
        create_all_shift_tasks()
    # корректировка разрешённых полей
    else:
        tasks = []
        # разрешённые поля
        allowed_fields = ('ws_number', 'norm_tech', 'doers_tech', 'draw_filename')
        for data in data_list:
            allowed_data = dict()
            for field in allowed_fields:
                allowed_data[field] = data.pop(field)
            try:
                task = ShiftTask.objects.get(**data)
                for key, value in allowed_data.items():
                    setattr(task, key, value)
                tasks.append(task)
            except ShiftTask.DoesNotExist:
                is_uploaded = False
        else:
            ShiftTask.objects.bulk_update(tasks, allowed_fields)
    return is_uploaded


def get_excel_data(data: Dict, exel_file: str, excel_list: str) -> List:  # TODO при рефакторинге перенести в services
    """
    Получение данных из excel
    :param data:
    :param exel_file:
    :param excel_list:
    :return:
    """
    ex_wb = openpyxl.load_workbook(exel_file, data_only=True)
    excel_list = excel_list.strip()
    ex_sh = ex_wb[excel_list.strip()]

    # шаблон номера операции
    op_number_template = r'\d\d.\d\d*'

    # определение максимальной строки для чтения
    max_read_row = 1
    for row in ex_sh.iter_rows(min_row=2, min_col=1, max_row=ex_sh.max_row,
                               max_col=2, values_only=True):
        max_read_row += 1
        if 'итого' in str(row[1]).lower():
            break

    data_list = []
    for i, row in enumerate(ex_sh.iter_rows(min_row=1, min_col=1, max_row=max_read_row,
                                            max_col=ex_sh.max_column, values_only=True)):
        if re.fullmatch(op_number_template, str(row[0])):
            data.update(
                {
                    'excel_list_name': f'{excel_list}-{i}',  # инкремент для гарантированно уникальной записи
                    'op_number': row[0],
                    'op_name': row[1],
                    'ws_name': row[2],
                    'op_name_full': row[1] + '-' + row[2],
                    'ws_number': str(row[3]),
                    'norm_tech': row[11],
                    'draw_filename': row[15],
                }
            )
            previous_op_number = str(row[0])  # обработка объединенных ячеек имени операции
            previous_op_name = row[1]
        elif row[0] is None and row[2] is not None:  # если объединённая ячейка
            data.update(
                {
                    'excel_list_name': f'{excel_list}-{i}',
                    'op_number': previous_op_number,
                    'op_name': previous_op_name,
                    'ws_name': row[2],
                    'op_name_full': previous_op_name + '-' + row[2],
                    'ws_number': str(row[3]),
                    'norm_tech': row[11],
                    'draw_filename': row[15],
                }
            )
        else:
            continue
        data_list.append(copy.copy(data))
    return data_list


if __name__ == '__main__':
    pass
