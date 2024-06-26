from pathlib import Path
from datetime import datetime
from typing import Any

from django.utils.timezone import make_naive
from django.db.models import QuerySet

from openpyxl import Workbook

from m_logger_settings import logger
from omzit_terminal.settings import BASE_DIR

from controller.models import DefectAct
from controller.utils.utils import get_model_verbose_names

# столбцы, выводимые в отчет
report_fields =[
    "datetime_fail", "act_number", "workshop", "operation", "processing_object",
    "control_object", "quantity", "inconsistencies", "remark", "tech_service",
    "tech_solution", "cause", "fixable", "fixing_time", "fio_failer", "master_finish_wp",
]

def defect_record_to_dict(record: DefectAct, fields: list[str]) -> dict[str, Any]:
    """
    Переводит запись из таблицы DefeectAct в словарь.
    Отдельно обрабатывает поля 'cause' и 'fixable'.
    В поле 'cause' содержатся сокращенные записи по чьей вине допущен брак. Чтобы получить человекочитаемую
    запись, надо преобразовать кортеж DefectAct.CHOICES в словарь и по ключу добратьс к полной фразе.
    'fixable' - можно починить: False, True и None. Для отчета в экселе нужно переводить на русский.
    @param record: запись из таблицы DefeectAct
    @param fields: поля, которые должны присутствовать в словаре
    @return:
    """
    res = {}
    choices_dict = dict(record.CHOICES)
    for key in fields:
        res[key] = getattr(record, key)
        if key == 'cause':
            res[key] = choices_dict.get(res[key], "")
        elif key == 'fixable':
            res[key] = [v for k, v in {True: 'да', False: "нет", None: ""}.items() if k == res[key]][0]
    return res

def queryset_to_dict(qs: QuerySet, fields) -> list[dict[str, Any]]:
    """
    Преобразует QuerySet в список словарей. Каждый словарь, это строка таблицы,
    а ключ словаря указывает на конкретную ячейку в строке.
    """
    table_dict = []
    for rec in qs:
        rec_dict = defect_record_to_dict(rec, fields)
        table_dict.append(rec_dict)
    return table_dict



def create_report():
    """
    Создает очтет обактах о браке. Копироует все записи из таблицы DefectAct
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Акты о браке"

    qs = DefectAct.objects.all()
    qs = queryset_to_dict(qs, report_fields)
    verbose_header = get_model_verbose_names(DefectAct)

    for i, column in enumerate(report_fields):
        ws.cell(row=1, column=i + 1).value = verbose_header[column]

    for i, row in enumerate(qs):
        for j, key in enumerate(row):
            cell = ws.cell(row=i + 2, column=j + 1)
            try:
                row[key] = make_naive(row[key])
            except Exception:
                pass
            try:
                cell.value = row[key]
            except ValueError as e:
                logger.error(f"Ошибка при конвертации значения {row[key]} по ключу {key}")
                logger.exception(e)
                cell.value = "error"

    x = datetime.now()
    filename = f"Акты о браке {x.day:02d}_{x.month:02d}_{x.year:04d} {x.hour:02d}_{x.minute:02d}_{x.second:02d}.xlsx"
    exel_file_dst = Path(BASE_DIR).joinpath("xlsx")
    abspath = exel_file_dst.joinpath(filename)
    wb.save(abspath)
    return abspath
