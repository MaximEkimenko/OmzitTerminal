import datetime
from typing import List

import psycopg2
from omzit_terminal.worker.services.db_config import host, dbname, user, password
import openpyxl


def select_master_call(ws_number: str) -> list or None:
    """
    Функция делает выборку из базы по РЦ и master_called = не было и формирует список сообщений из результатов запроса.
    Меняет master_called на вызван
    :param ws_number: номер рабочего центра
    :return: список сообщений для мастера
    """
    messages_to_master = []  # список сообщений для мастера
    try:
        # подключение к БД
        con = psycopg2.connect(dbname=dbname, user=user, password=password, host=host)
        con.autocommit = True
        # запрос на все статусы ожидания мастера
        select_query = f"""SELECT id, model_name, "order", op_number, op_name_full, fio_doer
                        FROM shift_task
                        WHERE st_status='ожидание мастера' AND 
                        ws_number = '{ws_number}' AND
                        master_called = 'не было'                
                        """
        try:
            with con.cursor() as cur:
                cur.execute(select_query)
                con.commit()
                shift_tasks = cur.fetchall()
                for task in shift_tasks:
                    # print(task)
                    message_to_master = (f"Вас ожидают на Т{ws_number}. Номер СЗ: {task[0]}. Заказ: {task[1]}. "
                                         f"Изделие: {task[2]}. Операция: {task[3]} {task[4]}. "
                                         f"Исполнители: {task[5]}")
                    messages_to_master.append(message_to_master)
        except Exception as e:
            print(e, 'ошибка выборке')
        if not messages_to_master:
            return None
        else:  # обновление переменной факта вызова мастера
            print('Обновление статуса')
            update_query = f"""UPDATE shift_task SET master_called = 'вызван'
                                        WHERE st_status='ожидание мастера' AND 
                                        ws_number = '{ws_number}' AND
                                        master_called = 'не было'; 
                            """
            try:
                with con.cursor() as cur:
                    cur.execute(update_query)
                    con.commit()
            except Exception as e:
                print(e, 'ошибка обновления')
    except Exception as e:
        print('Ошибка подключения к базе', e)
    finally:
        con.close()
    # print(messages_to_master)
    return messages_to_master


def ws_list_get(status: str) -> tuple:
    """
    Получение tuple уникальных РЦ запросом из базы со статусом status
    :param status: строка статуса
    :return: tuple РЦ
    """
    ws_list = set()  # список сообщений для мастера
    try:
        # подключение к БД
        con = psycopg2.connect(dbname=dbname, user=user, password=password, host=host)
        con.autocommit = True
        # запрос на все статусы ожидания мастера получение списка контролёров
        select_query = f"""SELECT ws_number
                            FROM shift_task
                            WHERE st_status='{status}';"""
        try:
            with con.cursor() as cur:
                cur.execute(select_query)
                con.commit()
                for shift_task in cur.fetchall():
                    ws_list.add(shift_task[0])
        except Exception as e:
            print(e, 'ошибка выборке')
    except Exception as e:
        print('Ошибка подключения к базе', e)
    finally:
        con.close()
    ws_list = tuple(sorted(ws_list))
    # print(ws_list)
    return ws_list


def st_list_get(ws_number: str) -> tuple:
    """
    Получение tuple СЗ с РЦ запросом из базы со статусом status
    :return: tuple РЦ
    """
    st_set = set()  # список сообщений для мастера
    try:
        # подключение к БД
        con = psycopg2.connect(dbname=dbname, user=user, password=password, host=host)
        con.autocommit = True
        # запрос на РЦ со статусом ожидание контролёра
        select_query = f"""SELECT id, ws_number, op_name_full
                            FROM shift_task
                            WHERE st_status='ожидание контролёра' AND
                            ws_number = '{ws_number}';
                            """
        try:
            with con.cursor() as cur:
                cur.execute(select_query)
                all_tasks = list(cur.fetchall())
        except Exception as e:
            print(e, 'ошибка выборке')
        for shift_task in all_tasks:
            st = f"№ {shift_task[0]} | T{shift_task[1]} | {shift_task[2]}"  # форматирование строки
            # print('ST-------', st)
            st_set.add(st)
    except Exception as e:
        print('Ошибка подключения к базе', e)
    finally:
        con.close()
    st_set = tuple(sorted(st_set))
    print('ST_SET----', st_set)
    return st_set


def master_id_get(ws_number: str = None, st_id: str = None) -> tuple:
    """
    Получает id мастера и ws_number по РЦ ws_number или id записи
    :param st_id: id записи в БД
    :param ws_number: номер РЦ
    :return: id мастера
    """
    if ws_number:
        query_field = 'ws_number'
        query_var = ws_number
    elif st_id:
        query_field = 'id'
        query_var = st_id
    print('ws_number=', ws_number, 'st_id=', st_id)
    try:
        # подключение к БД
        con = psycopg2.connect(dbname=dbname, user=user, password=password, host=host)
        con.autocommit = True
        # запрос на РЦ со статусом ожидание контролёра
        select_query = f"""SELECT master_finish_wp, ws_number
                                FROM shift_task
                                WHERE st_status='ожидание контролёра' AND
                                {query_field} = '{query_var}';"""
        try:
            with con.cursor() as cur:
                cur.execute(select_query)
                con.commit()
                master_id, ws_result_number = cur.fetchone()

        except Exception as e:
            print(e, 'ошибка выборке ПО МАСТЕРУ')
    except Exception as e:
        print('Ошибка подключения к базе', e)
    finally:
        con.close()
    print(master_id, ws_result_number)
    return master_id, ws_result_number


def status_change_to_otk(ws_number: str, initiator_id: str) -> None:
    """
    Изменение данных СЗ при вызове контролёра
    :param ws_number: номер РЦ
    :param initiator_id: telegram_id мастера
    :return: None
    """
    try:
        # подключение к БД
        con = psycopg2.connect(dbname=dbname, user=user, password=password, host=host)
        con.autocommit = True
        # запрос на корректировку БД
        update_query = f"""UPDATE shift_task SET st_status = 'ожидание контролёра',
                                        master_finish_wp = '{initiator_id}', 
                                        datetime_otk_call = '{datetime.datetime.now()}'
                                        WHERE st_status='ожидание мастера' AND 
                                        ws_number = '{ws_number}'"""
        try:
            with con.cursor() as cur:
                cur.execute(update_query)
                con.commit()
        except Exception as e:
            print(e, 'ошибка выборке')
    except Exception as e:
        print('Ошибка подключения к базе', e)
    finally:
        con.close()


def lines_count(ws_number: str) -> int:
    """
    Количество записей по ws_number
    :param ws_number:
    :return:
    """
    try:
        # подключение к БД
        con = psycopg2.connect(dbname=dbname, user=user, password=password, host=host)
        con.autocommit = True
        # запрос на корректировку БД
        count_query = f"""select COUNT(id) from shift_task where ws_number='{ws_number}' AND
                           st_status='ожидание контролёра'"""
        try:
            with con.cursor() as cur:
                cur.execute(count_query)
                con.commit()
                ws_count = cur.fetchall()
        except Exception as e:
            print(e, 'ошибка выборке')
        print(ws_count[0])
    except Exception as e:
        print('Ошибка подключения к базе', e)
    finally:
        con.close()
    return ws_count[0][0]


def control_man_id_set(ws_number, controlman_id):
    try:
        # подключение к БД
        con = psycopg2.connect(dbname=dbname, user=user, password=password, host=host)
        con.autocommit = True
        # запрос на корректировку БД
        update_query = f"""UPDATE shift_task SET otk_answer = '{controlman_id}', 
                                        datetime_otk_answer = '{datetime.datetime.now()}'
                                        WHERE st_status='ожидание контролёра' AND 
                                        ws_number = '{ws_number}'"""
        try:
            with con.cursor() as cur:
                cur.execute(update_query)
                con.commit()
        except Exception as e:
            print(e, 'ошибка выборке')
    except Exception as e:
        print('Ошибка подключения к базе', e)
    finally:
        con.close()


def decision_data_set(st_id, controlman_id, decision):
    try:
        # подключение к БД
        con = psycopg2.connect(dbname=dbname, user=user, password=password, host=host)
        con.autocommit = True
        # запрос на корректировку БД

        if decision == 'брак':
            update_query = f"""UPDATE shift_task SET otk_decision = '{controlman_id}',
                                                        st_status = '{decision}',
                                                        decision_time = '{datetime.datetime.now()}',
                                                        is_fail = 'True',
                                                        datetime_fail = '{datetime.datetime.now()}',
                                                        fio_failer = 
                                                        (SELECT fio_doer FROM shift_task WHERE id='{st_id}'),
                                                        job_duration = job_duration + ('{datetime.datetime.now()}' - datetime_job_resume)                                                      
                                                        WHERE id='{st_id}'"""
        else:
            update_query = f"""UPDATE shift_task SET otk_decision = '{controlman_id}',
                                            st_status = '{decision}',
                                            decision_time = '{datetime.datetime.now()}',
                                            job_duration = job_duration + ('{datetime.datetime.now()}' - datetime_job_resume)
                                            WHERE id='{st_id}'"""
        try:
            with con.cursor() as cur:
                cur.execute(update_query)
                con.commit()
        except Exception as e:
            print(e, 'ошибка выборке')
    except Exception as e:
        print('Ошибка подключения к базе', e)
    finally:
        con.close()


def indexes_calculation(st_id):
    """
    Получение данных из БД и расчёт необходимых показателей
    :param st_id: если = *, то выборка будет по всем записям
    :return:
    """
    results = dict()
    try:
        # подключение к БД
        con = psycopg2.connect(dbname=dbname, user=user, password=password, host=host)
        con.autocommit = True
        if st_id == '*':
            select_query = f"""SELECT * FROM shift_task"""  # получение всех данных
        else:
            select_query = f"""SELECT * FROM shift_task WHERE id='{st_id}'"""  # получение данных по id записи
        try:
            with con.cursor() as cur:
                cur.execute(select_query)
                con.commit()
                # print(cur.fetchall())
                # назначение переменных
                for line in cur.fetchall():
                    results['id'] = line[0]
                    results['workshop'] = line[1]
                    results['model_name'] = line[2]
                    results['datetime_done'] = line[3]
                    results['op_number'] = line[4]
                    results['op_name_full'] = line[7]
                    results['ws_number'] = line[8]
                    results['norm_tech'] = line[9]
                    results['order'] = line[12]
                    results['datetime_job_start'] = line[16]
                    results['decision_time'] = line[23]

        except Exception as e:
            print(e, 'ошибка выборке')
    except Exception as e:
        print('Ошибка подключения к базе', e)
    finally:
        con.close()
    print(results)
    timedelta = results['decision_time'] - results['datetime_job_start']
    print(type(timedelta))
    print(timedelta.total_seconds() / 60 / 60)


#
# a = datetime.timedelta()
# a.total_seconds()


def doers_update(excel_file: str) -> None:
    """
    Функция обновления таблицы doers (либо первоначального заполнения) из excel файла
    :param excel_file:
    :return:
    """
    wb = openpyxl.load_workbook(excel_file)
    sh = wb.active
    try:
        # подключение к БД
        con = psycopg2.connect(dbname=dbname, user=user, password=password, host=host)
        con.autocommit = True
        # запрос на корректировку БД
        for row in sh.iter_rows(min_row=2, max_row=sh.max_row, min_col=1, max_col=sh.max_column, values_only=True):
            if row[0] is not None:
                print(row[0])
                insert_query = f"""INSERT INTO doers (doers) VALUES ('{row[0]}')"""
                try:
                    with con.cursor() as cur:
                        cur.execute(insert_query)
                        con.commit()
                except Exception as e:
                    print(e, 'ошибка в запросе')
    except Exception as e:
        print('Ошибка подключения к базе', e)
    finally:
        con.close()

# TODO ЗАКОНСЕРВИРОВАНО Функционал простоев
# def confirm_downtime(shift_task_number, master_fio, description) -> None:
#     """
#     Подтверждение простоя
#     """
#     try:
#         # подключение к БД
#         con = psycopg2.connect(dbname=dbname, user=user, password=password, host=host)
#         con.autocommit = True
#         update_query = f""" UPDATE shift_task
#                             SET st_status = 'простой',
#                                 job_duration = job_duration + ('{datetime.datetime.now()}' - datetime_job_resume)
#                             WHERE id='{shift_task_number}';
#
#                             UPDATE downtime
#                             SET status = 'подтверждено',
#                                 description = '{description}',
#                                 datetime_start = '{datetime.datetime.now()}',
#                                 datetime_decision = '{datetime.datetime.now()}',
#                                 master_decision_fio = '{master_fio}'
#                             WHERE shift_task_id = '{shift_task_number}'
#                         """
#         try:
#             with con.cursor() as cur:
#                 cur.execute(update_query)
#                 con.commit()
#         except Exception as e:
#             print(e, 'ошибка выборке')
#     except Exception as e:
#         print('Ошибка подключения к базе', e)
#     finally:
#         con.close()
#
#
# def reject_downtime(shift_task_number, master_fio, description) -> None:
#     """
#     Отклонение простоя
#     """
#     try:
#         # подключение к БД
#         con = psycopg2.connect(dbname=dbname, user=user, password=password, host=host)
#         con.autocommit = True
#         update_query = f""" UPDATE downtime
#                             SET status = 'отклонено',
#                                 description = '{description}',
#                                 datetime_decision = '{datetime.datetime.now()}',
#                                 master_decision_fio = '{master_fio}'
#                             WHERE shift_task_id = '{shift_task_number}'
#                         """
#         try:
#             with con.cursor() as cur:
#                 cur.execute(update_query)
#                 con.commit()
#         except Exception as e:
#             print(e, 'ошибка выборке')
#     except Exception as e:
#         print('Ошибка подключения к базе', e)
#     finally:
#         con.close()


if __name__ == '__main__':
    # select_master_call("13")
    # ws_list_get('')
    # status_change_to_otk('109', 123)
    # st_list_get('109')
    # print(st_list_get('109')[0][2:5], len(st_list_get('109')[0][2:5]))
    # master_id_get(st_id='275')
    # master_id_get(ws_number='109')
    # decision_data_set('264', '123', '!!!!!')
    # indexes_calculation('274')
    # fio_file = r'D:\АСУП\Python\Projects\ARC\GUI\Табель\База ФИО.xlsx'
    # status_change_to_otk('13', '1238658905')
    # doers_update(fio_file)
    pass
